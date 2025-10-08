"""Worksheet component base classes and implementations.

Provides component architecture for building Excel worksheets with
consistent styling, layout, and functionality.
"""

from abc import ABC, abstractmethod
from typing import Dict, Any, List, Optional

import structlog
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = structlog.get_logger(__name__)


class WorksheetComponent(ABC):
    """
    Base class for worksheet components.

    Each component is responsible for generating a specific aspect of the
    Excel template, such as task lists, Gantt charts, or summary sheets.
    """

    @abstractmethod
    def generate(self, worksheet: Worksheet, config: Dict[str, Any]) -> None:
        """
        Generate worksheet content.

        Args:
            worksheet: openpyxl Worksheet instance to populate
            config: Configuration dictionary with component-specific settings
        """
        pass

    @property
    @abstractmethod
    def name(self) -> str:
        """Component name for logging and identification."""
        pass


class TaskListComponent(WorksheetComponent):
    """Component for generating task list structure."""

    @property
    def name(self) -> str:
        return "TaskList"

    def generate(self, worksheet: Worksheet, config: Dict[str, Any]) -> None:
        """
        Generate task list structure with headers and formatting.

        Args:
            worksheet: Worksheet to populate
            config: Configuration with task list settings
        """
        headers = config.get(
            "headers",
            [
                "Task ID",
                "Task Name",
                "Duration",
                "Start Date",
                "End Date",
                "Dependencies",
                "Sprint",
                "Status",
                "Owner",
            ],
        )

        # Apply header styling
        self._apply_header_styling(worksheet, headers)

        # Set column widths
        self._set_column_widths(worksheet, headers)

        logger.debug("Task list component generated", headers_count=len(headers))

    def _apply_header_styling(self, worksheet: Worksheet, headers: List[str]) -> None:
        """Apply consistent styling to header row."""
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _set_column_widths(self, worksheet: Worksheet, headers: List[str]) -> None:
        """Set appropriate column widths based on content."""
        # Default widths for common columns
        width_map = {
            "Task ID": 10,
            "Task Name": 30,
            "Duration": 12,
            "Start Date": 12,
            "End Date": 12,
            "Dependencies": 20,
            "Sprint": 10,
            "Status": 12,
            "Owner": 15,
        }

        for col_idx, header in enumerate(headers, start=1):
            width = width_map.get(header, 15)  # Default 15 if not in map
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width


class GanttChartComponent(WorksheetComponent):
    """Component for generating Gantt chart visualization area."""

    @property
    def name(self) -> str:
        return "GanttChart"

    def generate(self, worksheet: Worksheet, config: Dict[str, Any]) -> None:
        """
        Generate Gantt chart area with timeline headers.

        Args:
            worksheet: Worksheet to populate
            config: Configuration with Gantt chart settings
        """
        start_column = config.get("start_column", 10)
        num_weeks = config.get("num_weeks", 12)

        # Add timeline header
        self._add_timeline_header(worksheet, start_column, num_weeks)

        logger.debug(
            "Gantt chart component generated",
            start_column=start_column,
            num_weeks=num_weeks,
        )

    def _add_timeline_header(
        self, worksheet: Worksheet, start_column: int, num_weeks: int
    ) -> None:
        """Add timeline header for Gantt chart."""
        timeline_fill = PatternFill(
            start_color="D0CECE", end_color="D0CECE", fill_type="solid"
        )
        timeline_font = Font(bold=True, size=9)
        timeline_alignment = Alignment(horizontal="center", vertical="center")

        for week in range(num_weeks):
            col_idx = start_column + week
            cell = worksheet.cell(row=1, column=col_idx, value=f"Week {week + 1}")
            cell.fill = timeline_fill
            cell.font = timeline_font
            cell.alignment = timeline_alignment

            # Set column width for Gantt columns
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 3


class MetadataComponent(WorksheetComponent):
    """Component for generating metadata and configuration sections."""

    @property
    def name(self) -> str:
        return "Metadata"

    def generate(self, worksheet: Worksheet, config: Dict[str, Any]) -> None:
        """
        Generate metadata section with project information.

        Args:
            worksheet: Worksheet to populate
            config: Configuration with metadata
        """
        # Add project information section
        project_name = config.get("project_name", "Unnamed Project")
        sprint_pattern = config.get("sprint_pattern", "YY.Q.#")

        # Create metadata area at top of sheet (if desired)
        # This is optional - main metadata is in hidden _SYNC_META sheet

        logger.debug("Metadata component generated", project_name=project_name)
