"""
Excel Template Engine for SprintForge.

Component-based architecture for generating sophisticated Excel spreadsheets
with project management formulas, Gantt charts, and sync metadata.
"""

import hashlib
import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Dict, Any, Optional, List
from pathlib import Path

import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.excel.components.worksheets import WorksheetComponent
from app.excel.components.formulas import FormulaTemplate

logger = structlog.get_logger(__name__)


class ProjectConfig:
    """Project configuration for Excel generation."""

    def __init__(
        self,
        project_id: str,
        project_name: str,
        sprint_pattern: str = "YY.Q.#",
        features: Optional[Dict[str, bool]] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ):
        self.project_id = project_id
        self.project_name = project_name
        self.sprint_pattern = sprint_pattern
        self.features = features or {}
        self.metadata = metadata or {}


class ExcelTemplateEngine:
    """
    Main engine for generating Excel templates from project configurations.

    Uses a component-based architecture where each aspect of the Excel file
    (worksheets, formulas, styles) is handled by specialized components.

    Example:
        >>> engine = ExcelTemplateEngine()
        >>> config = ProjectConfig(
        ...     project_id="proj_123",
        ...     project_name="My Project",
        ...     sprint_pattern="YY.Q.#"
        ... )
        >>> excel_bytes = engine.generate_template(config)
    """

    def __init__(self):
        """Initialize the Excel template engine with registered components."""
        self.components: List[WorksheetComponent] = []
        self.formula_templates = FormulaTemplate()
        logger.info("Excel template engine initialized")

    def register_component(self, component: WorksheetComponent) -> None:
        """
        Register a worksheet component for template generation.

        Args:
            component: WorksheetComponent instance to register
        """
        self.components.append(component)
        logger.debug("Component registered", component=component.__class__.__name__)

    def generate_template(self, config: ProjectConfig) -> bytes:
        """
        Generate an Excel template from project configuration.

        Args:
            config: ProjectConfig with project details and settings

        Returns:
            bytes: Excel file content as bytes

        Raises:
            ValueError: If configuration is invalid
        """
        logger.info(
            "Starting Excel template generation",
            project_id=config.project_id,
            project_name=config.project_name,
        )

        try:
            # Create new workbook
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet

            # Generate main worksheet
            self._create_main_worksheet(workbook, config)

            # Add sync metadata worksheet (hidden)
            self._create_sync_metadata(workbook, config)

            # Save to bytes
            excel_bytes = self._save_to_bytes(workbook)

            logger.info(
                "Excel template generated successfully",
                project_id=config.project_id,
                size_bytes=len(excel_bytes),
            )

            return excel_bytes

        except Exception as e:
            logger.error(
                "Excel template generation failed",
                project_id=config.project_id,
                error=str(e),
            )
            raise

    def _create_main_worksheet(self, workbook: Workbook, config: ProjectConfig) -> None:
        """
        Create the main project management worksheet.

        Args:
            workbook: openpyxl Workbook instance
            config: Project configuration
        """
        ws = workbook.create_sheet(title="Project Plan")

        # Set up header row
        headers = [
            "Task ID",
            "Task Name",
            "Duration (days)",
            "Start Date",
            "End Date",
            "Dependencies",
            "Sprint",
            "Status",
            "Owner",
        ]

        # Write headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Set column widths
        column_widths = [10, 30, 15, 12, 12, 20, 10, 12, 15]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add sample task row (for template demonstration)
        ws.cell(row=2, column=1, value="T001")
        ws.cell(row=2, column=2, value="Sample Task")
        ws.cell(row=2, column=3, value=5)
        ws.cell(row=2, column=4, value=datetime.now().date())
        ws.cell(row=2, column=8, value="Not Started")

        logger.debug("Main worksheet created", sheet_name=ws.title)

    def _create_sync_metadata(self, workbook: Workbook, config: ProjectConfig) -> None:
        """
        Create hidden metadata worksheet for sync functionality.

        This worksheet stores project metadata needed for two-way sync between
        Excel and the server, including project ID, version, and checksums.

        Args:
            workbook: openpyxl Workbook instance
            config: Project configuration
        """
        ws = workbook.create_sheet(title="_SYNC_META")

        # Hide the metadata sheet
        ws.sheet_state = "hidden"

        # Create metadata structure
        metadata = {
            "project_id": config.project_id,
            "project_name": config.project_name,
            "version": "1.0.0",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "sprint_pattern": config.sprint_pattern,
            "features": config.features,
            "checksum": self._calculate_checksum(config),
        }

        # Write metadata as JSON in first cell (A1)
        ws["A1"] = json.dumps(metadata, indent=2)
        ws["A2"] = "DO NOT MODIFY - Required for sync functionality"

        # Style metadata cell
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["A2"].font = Font(italic=True, color="999999")

        logger.debug("Sync metadata worksheet created", project_id=config.project_id)

    def _calculate_checksum(self, config: ProjectConfig) -> str:
        """
        Calculate checksum for project configuration.

        Args:
            config: Project configuration

        Returns:
            str: SHA-256 checksum hex string
        """
        config_str = json.dumps(
            {
                "project_id": config.project_id,
                "project_name": config.project_name,
                "sprint_pattern": config.sprint_pattern,
            },
            sort_keys=True,
        )
        return hashlib.sha256(config_str.encode()).hexdigest()

    def _save_to_bytes(self, workbook: Workbook) -> bytes:
        """
        Save workbook to bytes.

        Args:
            workbook: openpyxl Workbook instance

        Returns:
            bytes: Excel file content
        """
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def load_metadata_from_excel(self, excel_bytes: bytes) -> Dict[str, Any]:
        """
        Extract sync metadata from an Excel file.

        Args:
            excel_bytes: Excel file content as bytes

        Returns:
            Dict containing metadata from _SYNC_META sheet

        Raises:
            ValueError: If metadata sheet is missing or invalid
        """
        from openpyxl import load_workbook

        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        if "_SYNC_META" not in workbook.sheetnames:
            raise ValueError("Missing _SYNC_META worksheet - file not generated by SprintForge")

        meta_sheet = workbook["_SYNC_META"]
        metadata_json = meta_sheet["A1"].value

        if not metadata_json:
            raise ValueError("Empty metadata in _SYNC_META worksheet")

        try:
            metadata = json.loads(metadata_json)
            logger.info("Metadata loaded from Excel", project_id=metadata.get("project_id"))
            return metadata
        except json.JSONDecodeError as e:
            logger.error("Invalid metadata JSON", error=str(e))
            raise ValueError(f"Invalid metadata format: {e}")
