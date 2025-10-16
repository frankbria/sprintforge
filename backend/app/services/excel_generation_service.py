"""
ExcelGenerationService - Enhanced Excel generation for Monte Carlo simulations.

Creates Excel templates with PERT formula columns, Monte Carlo Results sheet,
and Quick Simulation functionality for testing and validation.
"""

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.simulation_service import SimulationResult


class ExcelGenerationService:
    """
    Service for generating enhanced Excel templates with Monte Carlo capabilities.

    Creates Excel workbooks with:
    - Task List sheet with PERT formula columns
    - Monte Carlo Results sheet for simulation output
    - Quick Simulation sheet with 100 pre-filled sample tasks
    - Professional formatting (headers, borders, colors)

    Usage:
        service = ExcelGenerationService()
        workbook = service.create_template_workbook(
            project_name="My Project",
            include_sample_data=True
        )
        service.apply_formatting(workbook)
        excel_bytes = service.save_workbook_to_bytes(workbook)
    """

    # Color constants for formatting
    HEADER_FILL = PatternFill(
        start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
    )
    PERT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    BORDER_STYLE = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def create_template_workbook(
        self, project_name: str = "New Project", include_sample_data: bool = False
    ) -> Workbook:
        """
        Create Excel template with enhanced PERT columns.

        Args:
            project_name: Name of the project for the template
            include_sample_data: Whether to include sample task rows

        Returns:
            Workbook with Task List sheet configured
        """
        workbook = Workbook()

        # Remove default sheet and create Task List
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        task_sheet = workbook.create_sheet("Task List", 0)

        # Create headers
        headers = [
            "Task ID",
            "Task Name",
            "Optimistic Duration",
            "Most Likely Duration",
            "Pessimistic Duration",
            "PERT Mean",
            "Dependencies",
            "Notes",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = task_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add sample data if requested
        if include_sample_data:
            sample_tasks = self._generate_sample_tasks(count=5)
            self._populate_task_sheet(task_sheet, sample_tasks, start_row=2)

        # Set column widths
        task_sheet.column_dimensions["A"].width = 12
        task_sheet.column_dimensions["B"].width = 25
        task_sheet.column_dimensions["C"].width = 20
        task_sheet.column_dimensions["D"].width = 22
        task_sheet.column_dimensions["E"].width = 22
        task_sheet.column_dimensions["F"].width = 15
        task_sheet.column_dimensions["G"].width = 20
        task_sheet.column_dimensions["H"].width = 30

        # Freeze header row
        task_sheet.freeze_panes = "A2"

        return workbook

    def add_monte_carlo_results_sheet(
        self,
        workbook: Workbook,
        simulation_result: SimulationResult,
        tasks: List[Dict[str, Any]],
        critical_path: List[str],
    ) -> None:
        """
        Add Monte Carlo results to existing workbook.

        Args:
            workbook: Workbook to add results sheet to
            simulation_result: SimulationResult from Monte Carlo simulation
            tasks: List of task dictionaries with task information
            critical_path: List of task IDs forming the critical path
        """
        # Create or get Monte Carlo Results sheet
        if "Monte Carlo Results" in workbook.sheetnames:
            del workbook["Monte Carlo Results"]

        results_sheet = workbook.create_sheet("Monte Carlo Results")

        # Create headers
        headers = [
            "Simulation Date/Time",
            "Iterations",
            "Mean Duration (days)",
            "Median Duration (P50)",
            "Standard Deviation",
        ]

        # Add percentile headers
        percentiles = sorted(simulation_result.confidence_intervals.keys())
        for p in percentiles:
            headers.append(f"P{p}")

        headers.extend(["Task Count", "Critical Path"])

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = results_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        row_data = [
            simulation_result.simulation_date.strftime("%Y-%m-%d %H:%M:%S"),
            simulation_result.iterations_run,
            round(simulation_result.mean_duration, 2),
            round(simulation_result.median_duration, 2),
            round(simulation_result.std_deviation, 2),
        ]

        # Add percentile values
        for p in percentiles:
            row_data.append(round(simulation_result.confidence_intervals[p], 2))

        row_data.append(len(tasks))
        row_data.append(", ".join(critical_path))

        for col_idx, value in enumerate(row_data, start=1):
            results_sheet.cell(row=2, column=col_idx, value=value)

        # Set column widths
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            results_sheet.column_dimensions[col_letter].width = 18

        # Freeze header row
        results_sheet.freeze_panes = "A2"

    def create_quick_simulation_sheet(self, workbook: Workbook) -> None:
        """
        Add Quick Simulation sheet with 100 sample tasks.

        Args:
            workbook: Workbook to add Quick Simulation sheet to
        """
        # Create or get Quick Simulation sheet
        if "Quick Simulation" in workbook.sheetnames:
            del workbook["Quick Simulation"]

        quick_sim_sheet = workbook.create_sheet("Quick Simulation")

        # Create headers (same as Task List)
        headers = [
            "Task ID",
            "Task Name",
            "Optimistic Duration",
            "Most Likely Duration",
            "Pessimistic Duration",
            "PERT Mean",
            "Dependencies",
            "Notes",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = quick_sim_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Generate 100 sample tasks
        sample_tasks = self._generate_sample_tasks(count=100)
        self._populate_task_sheet(quick_sim_sheet, sample_tasks, start_row=2)

        # Set column widths
        quick_sim_sheet.column_dimensions["A"].width = 12
        quick_sim_sheet.column_dimensions["B"].width = 25
        quick_sim_sheet.column_dimensions["C"].width = 20
        quick_sim_sheet.column_dimensions["D"].width = 22
        quick_sim_sheet.column_dimensions["E"].width = 22
        quick_sim_sheet.column_dimensions["F"].width = 15
        quick_sim_sheet.column_dimensions["G"].width = 20
        quick_sim_sheet.column_dimensions["H"].width = 30

        # Freeze header row
        quick_sim_sheet.freeze_panes = "A2"

    def apply_formatting(self, workbook: Workbook) -> None:
        """
        Apply professional Excel formatting.

        Applies:
        - Header formatting (bold, background color, freeze panes)
        - PERT Mean column highlighting (light green)
        - Number formatting (2 decimal places for durations)
        - Borders on all cells
        - Column width adjustments

        Args:
            workbook: Workbook to apply formatting to
        """
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Apply formatting to all populated cells
            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
            ):
                for cell in row:
                    # Apply borders
                    cell.border = self.BORDER_STYLE

                    # Apply header formatting
                    if cell.row == 1:
                        cell.font = Font(bold=True)
                        cell.fill = self.HEADER_FILL
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

                    # Apply PERT Mean column formatting (column F in Task List and Quick Simulation)
                    if (
                        sheet_name in ["Task List", "Quick Simulation"]
                        and cell.column == 6
                        and cell.row > 1
                    ):
                        cell.fill = self.PERT_FILL

                    # Apply number formatting to duration columns (C, D, E, F)
                    if cell.column in [3, 4, 5, 6] and cell.row > 1:
                        if cell.value is not None:
                            cell.number_format = "0.00"

    def save_workbook_to_bytes(self, workbook: Workbook) -> bytes:
        """
        Save workbook to bytes for download.

        Args:
            workbook: Workbook to save

        Returns:
            Excel file content as bytes
        """
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _generate_sample_tasks(self, count: int = 5) -> List[Dict[str, Any]]:
        """
        Generate sample tasks with varied durations and dependencies.

        Creates realistic test data with:
        - Varied task types (Planning, Development, Testing, etc.)
        - Realistic PERT estimates (optimistic < most_likely < pessimistic)
        - Dependency chains (30-40% of tasks have dependencies)
        - Durations ranging from 1 to 15 days

        Args:
            count: Number of sample tasks to generate (1-1000)

        Returns:
            List of task dictionaries with keys:
                task_id, task_name, optimistic, most_likely, pessimistic,
                dependencies, notes
        """
        import random

        tasks: List[Dict[str, Any]] = []
        task_types = [
            "Planning",
            "Design",
            "Development",
            "Testing",
            "Review",
            "Integration",
            "Deployment",
            "Documentation",
            "Research",
            "Analysis",
            "Architecture",
            "Implementation",
            "Quality Assurance",
            "Performance Testing",
            "Security Review",
            "User Acceptance",
            "Training",
            "Migration",
            "Optimization",
            "Monitoring",
        ]

        for i in range(1, count + 1):
            # Generate realistic duration estimates (PERT ordering)
            base_duration = random.uniform(1.0, 15.0)
            optimistic = round(base_duration * 0.7, 1)
            most_likely = round(base_duration, 1)
            pessimistic = round(base_duration * 1.5, 1)

            # Generate dependencies (30-40% of tasks have dependencies)
            dependencies = ""
            if i > 1 and random.random() < 0.4:
                # Depend on 1-2 previous tasks
                num_deps = random.randint(1, min(2, i - 1))
                dep_tasks = random.sample(range(1, i), num_deps)
                dependencies = ",".join([f"TASK-{d}" for d in dep_tasks])

            task_type = random.choice(task_types)

            task = {
                "task_id": f"TASK-{i}",
                "task_name": f"{task_type} {i}",
                "optimistic": optimistic,
                "most_likely": most_likely,
                "pessimistic": pessimistic,
                "dependencies": dependencies,
                "notes": "",
            }
            tasks.append(task)

        return tasks

    def _populate_task_sheet(
        self, sheet: Worksheet, tasks: List[Dict[str, Any]], start_row: int = 2
    ) -> None:
        """
        Populate task sheet with task data and PERT formulas.

        Args:
            sheet: Worksheet to populate
            tasks: List of task dictionaries
            start_row: Starting row number (usually 2, after header)
        """
        for idx, task in enumerate(tasks):
            row_num = start_row + idx

            # Write task data
            sheet.cell(row=row_num, column=1, value=task["task_id"])
            sheet.cell(row=row_num, column=2, value=task["task_name"])
            sheet.cell(row=row_num, column=3, value=task["optimistic"])
            sheet.cell(row=row_num, column=4, value=task["most_likely"])
            sheet.cell(row=row_num, column=5, value=task["pessimistic"])

            # PERT formula: =(C+4*D+E)/6
            pert_formula = f"=(C{row_num}+4*D{row_num}+E{row_num})/6"
            sheet.cell(row=row_num, column=6, value=pert_formula)

            sheet.cell(row=row_num, column=7, value=task["dependencies"])
            sheet.cell(row=row_num, column=8, value=task.get("notes", ""))
