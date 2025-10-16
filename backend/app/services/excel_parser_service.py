"""
Excel file parsing service for Monte Carlo simulation workflow.

Parses Excel files containing task data with PERT estimates, validates
structure and constraints, and converts to TaskDistributionInput for simulation.
"""

import io
from typing import Any, Dict, List, Optional, Set

import openpyxl  # type: ignore
from openpyxl.workbook import Workbook  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from pydantic import BaseModel, model_validator

from app.services.scheduler.monte_carlo import TaskDistributionInput


class ExcelParseError(Exception):
    """Raised when Excel parsing fails."""

    pass


class ParsedTask(BaseModel):
    """Task parsed from Excel file."""

    task_id: str
    task_name: str
    optimistic: float
    most_likely: float
    pessimistic: float
    dependencies: str = ""
    notes: Optional[str] = None

    @model_validator(mode="after")
    def validate_pert_order(self) -> "ParsedTask":
        """Ensure optimistic ≤ most_likely ≤ pessimistic."""
        if not (self.optimistic <= self.most_likely <= self.pessimistic):
            raise ValueError(
                f"Task {self.task_id}: Invalid PERT order "
                f"({self.optimistic}, {self.most_likely}, {self.pessimistic})"
            )
        return self


class ParsedExcelData(BaseModel):
    """Complete parsed Excel data."""

    tasks: List[ParsedTask]
    project_name: Optional[str] = None
    metadata: Dict[str, Any] = {}


# Column mapping for flexible header matching
COLUMN_MAPPINGS = {
    "task_id": ["Task ID", "ID", "Task", "TaskID"],
    "task_name": ["Task Name", "Name", "Description"],
    "optimistic": ["Optimistic", "Opt", "O", "Optimistic Duration"],
    "most_likely": ["Most Likely", "ML", "M", "Most Likely Duration"],
    "pessimistic": ["Pessimistic", "Pess", "P", "Pessimistic Duration"],
    "dependencies": ["Dependencies", "Deps", "Predecessors"],
    "notes": ["Notes", "Note", "Comments", "Comment"],
}


class ExcelParserService:
    """
    Service for parsing Excel files containing task data.

    Extracts task information including PERT estimates and dependencies,
    validates data structure and constraints, and converts to simulation input.
    """

    def parse_excel_file(self, file_bytes: bytes, filename: str) -> ParsedExcelData:
        """
        Parse Excel file and return structured task data.

        Args:
            file_bytes: Raw Excel file bytes
            filename: Name of the file (for error messages)

        Returns:
            ParsedExcelData with task list and metadata

        Raises:
            ExcelParseError: If parsing fails for any reason
        """
        # Step 1: Validate file is not empty
        if not file_bytes:
            raise ExcelParseError("Empty file")

        # Step 2: Load workbook
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise ExcelParseError(f"Not a valid Excel file: {e}")

        # Step 3: Find the correct sheet
        ws = self._find_task_sheet(wb)
        if ws is None:
            raise ExcelParseError("Could not find task data sheet")

        # Step 4: Parse headers and find column mappings
        column_indices = self._parse_headers(ws)

        # Step 5: Parse task rows
        tasks = self._parse_task_rows(ws, column_indices)

        # Step 6: Validate we have tasks
        if not tasks:
            raise ExcelParseError("No tasks found in Excel file")

        return ParsedExcelData(tasks=tasks)

    def _find_task_sheet(self, wb: Workbook) -> Optional[Worksheet]:
        """
        Find the sheet containing task data.

        Looks for "Task List" sheet first, otherwise uses active sheet.

        Args:
            wb: Workbook to search

        Returns:
            Worksheet containing task data
        """
        # Try to find "Task List" sheet
        for sheet_name in wb.sheetnames:
            if "task list" in sheet_name.lower():
                return wb[sheet_name]

        # Use active sheet as fallback
        return wb.active

    def _parse_headers(self, ws: Worksheet) -> Dict[str, int]:
        """
        Parse header row and map to column indices.

        Args:
            ws: Worksheet to parse

        Returns:
            Dictionary mapping field names to column indices

        Raises:
            ExcelParseError: If required columns are missing
        """
        # Find header row (first non-empty row)
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            if any(cell.value for cell in row):
                header_row = row
                break

        if header_row is None:
            raise ExcelParseError("Could not find header row")

        # Build column index mapping
        column_indices: Dict[str, int] = {}

        for col_idx, cell in enumerate(header_row):
            header_value = str(cell.value).strip() if cell.value else ""

            # Match against known column mappings
            for field_name, possible_headers in COLUMN_MAPPINGS.items():
                if header_value in possible_headers:
                    column_indices[field_name] = col_idx
                    break

        # Check for required columns
        required_fields = [
            "task_id",
            "task_name",
            "optimistic",
            "most_likely",
            "pessimistic",
        ]
        missing_fields = [f for f in required_fields if f not in column_indices]

        if missing_fields:
            raise ExcelParseError(
                f"Missing required columns: {', '.join(missing_fields)}"
            )

        return column_indices

    def _parse_task_rows(
        self, ws: Worksheet, column_indices: Dict[str, int]
    ) -> List[ParsedTask]:
        """
        Parse task rows from worksheet.

        Args:
            ws: Worksheet to parse
            column_indices: Mapping of field names to column indices

        Returns:
            List of parsed tasks

        Raises:
            ExcelParseError: If data validation fails
        """
        tasks: List[ParsedTask] = []

        # Start from row 2 (after headers)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue

            # Extract values
            try:
                task_id = self._get_cell_value(row, column_indices.get("task_id"), str)
                if not task_id:
                    continue  # Skip rows without task ID

                task_name = self._get_cell_value(
                    row, column_indices.get("task_name"), str
                )
                optimistic = self._get_cell_value(
                    row, column_indices.get("optimistic"), float
                )
                most_likely = self._get_cell_value(
                    row, column_indices.get("most_likely"), float
                )
                pessimistic = self._get_cell_value(
                    row, column_indices.get("pessimistic"), float
                )
                dependencies = self._get_cell_value(
                    row, column_indices.get("dependencies"), str, default=""
                )
                notes = self._get_cell_value(
                    row, column_indices.get("notes"), str, default=None
                )

                # Create task (will validate PERT order)
                task = ParsedTask(
                    task_id=task_id,
                    task_name=task_name,
                    optimistic=optimistic,
                    most_likely=most_likely,
                    pessimistic=pessimistic,
                    dependencies=dependencies,
                    notes=notes,
                )
                tasks.append(task)

            except ValueError as e:
                if "Invalid PERT order" in str(e):
                    raise ExcelParseError(str(e))
                raise ExcelParseError(f"Row {row_idx}: Invalid data type - {e}")
            except Exception as e:
                raise ExcelParseError(f"Row {row_idx}: Failed to parse - {e}")

        return tasks

    def _get_cell_value(
        self, row: tuple, col_idx: Optional[int], target_type: type, default: Any = None
    ) -> Any:
        """
        Extract and convert cell value to target type.

        Args:
            row: Row tuple from worksheet
            col_idx: Column index (or None if column not found)
            target_type: Type to convert to (str, float, int)
            default: Default value if cell is empty

        Returns:
            Converted value

        Raises:
            ValueError: If type conversion fails
        """
        if col_idx is None or col_idx >= len(row):
            return default

        cell_value = row[col_idx].value

        if cell_value is None or cell_value == "":
            return default

        # Convert to target type
        try:
            if target_type == str:
                return str(cell_value).strip()
            elif target_type == float:
                return float(cell_value)
            elif target_type == int:
                return int(cell_value)
            else:
                return cell_value
        except (ValueError, TypeError) as e:
            raise ValueError(
                f"Cannot convert '{cell_value}' to {target_type.__name__}: {e}"
            )

    def validate_task_structure(self, tasks: List[ParsedTask]) -> List[str]:
        """
        Validate task data structure and dependencies.

        Args:
            tasks: List of parsed tasks to validate

        Returns:
            List of error messages (empty if valid)
        """
        errors: List[str] = []

        # Build task ID set for validation
        task_ids = {task.task_id for task in tasks}

        # Check for unknown dependencies
        for task in tasks:
            if task.dependencies:
                deps = [d.strip() for d in task.dependencies.split(",") if d.strip()]
                for dep in deps:
                    if dep not in task_ids:
                        errors.append(
                            f"Task {task.task_id}: Unknown dependency "
                            f"'{dep}' does not exist"
                        )

        # Check for circular dependencies
        circular_error = self._detect_circular_dependencies(tasks)
        if circular_error:
            errors.append(circular_error)

        return errors

    def _detect_circular_dependencies(self, tasks: List[ParsedTask]) -> Optional[str]:
        """
        Detect circular dependencies using DFS.

        Args:
            tasks: List of tasks to check

        Returns:
            Error message if circular dependency found, None otherwise
        """
        # Build adjacency list
        graph: Dict[str, List[str]] = {}
        for task in tasks:
            if task.dependencies:
                deps = [d.strip() for d in task.dependencies.split(",") if d.strip()]
                graph[task.task_id] = deps
            else:
                graph[task.task_id] = []

        # DFS to detect cycles
        visited: Set[str] = set()
        rec_stack: Set[str] = set()

        def has_cycle(task_id: str) -> Optional[str]:
            visited.add(task_id)
            rec_stack.add(task_id)

            # Check all neighbors
            for neighbor in graph.get(task_id, []):
                if neighbor not in visited:
                    result = has_cycle(neighbor)
                    if result:
                        return result
                elif neighbor in rec_stack:
                    return f"Circular dependency detected involving task {neighbor}"

            rec_stack.remove(task_id)
            return None

        # Check all nodes
        for task_id in graph:
            if task_id not in visited:
                error = has_cycle(task_id)
                if error:
                    return error

        return None

    def convert_to_distribution_input(
        self, parsed_tasks: List[ParsedTask]
    ) -> List[TaskDistributionInput]:
        """
        Convert parsed tasks to TaskDistributionInput for simulation.

        Args:
            parsed_tasks: List of parsed tasks

        Returns:
            List of TaskDistributionInput objects
        """
        return [
            TaskDistributionInput(task_id=task.task_id, dependencies=task.dependencies)
            for task in parsed_tasks
        ]
