"""
Tests for ExcelParserService - Excel file parsing and validation.

TDD approach:
1. Write tests first (RED phase)
2. Implement minimal code to pass (GREEN phase)
3. Refactor while keeping tests green (REFACTOR phase)

Coverage target: ≥85%
Pass rate target: 100%
"""

import io
from datetime import date
from typing import List

import openpyxl
import pytest
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel_parser_service import (
    ExcelParseError,
    ExcelParserService,
    ParsedExcelData,
    ParsedTask,
)
from app.services.scheduler.monte_carlo import TaskDistributionInput


# Helper Functions
def create_valid_excel() -> bytes:
    """Create a valid Excel file with task data for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    # Headers
    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
            "Notes",
        ]
    )

    # Sample tasks
    ws.append(["T001", "Design", 3.0, 5.0, 8.0, "", "Initial design"])
    ws.append(["T002", "Development", 10.0, 15.0, 25.0, "T001", "Core features"])
    ws.append(["T003", "Testing", 5.0, 7.0, 10.0, "T002", "QA testing"])
    ws.append(["T004", "Deployment", 1.0, 2.0, 3.0, "T003", "Production deployment"])

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_with_missing_columns() -> bytes:
    """Create Excel file missing required columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    # Missing Optimistic, Most Likely, Pessimistic columns
    ws.append(["Task ID", "Task Name", "Dependencies"])
    ws.append(["T001", "Design", ""])
    ws.append(["T002", "Development", "T001"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_with_invalid_data_types() -> bytes:
    """Create Excel file with invalid data types."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    ws.append(["T001", "Design", "invalid", 5.0, 8.0, ""])  # Invalid optimistic
    ws.append(
        ["T002", "Development", 10.0, "invalid", 25.0, "T001"]
    )  # Invalid most_likely

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_with_pert_violation() -> bytes:
    """Create Excel file with PERT order violation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    ws.append(
        ["T001", "Design", 8.0, 5.0, 3.0, ""]
    )  # Violates optimistic ≤ most_likely ≤ pessimistic

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_with_circular_dependencies() -> bytes:
    """Create Excel file with circular dependencies."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    ws.append(["T001", "Design", 3.0, 5.0, 8.0, "T003"])  # Depends on T003
    ws.append(["T002", "Development", 10.0, 15.0, 25.0, "T001"])  # Depends on T001
    ws.append(
        ["T003", "Testing", 5.0, 7.0, 10.0, "T002"]
    )  # Depends on T002 -> circular

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_empty_excel() -> bytes:
    """Create Excel file with headers but no data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    # No data rows

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_with_alternative_headers() -> bytes:
    """Create Excel file with alternative header names."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    # Alternative header names
    ws.append(["ID", "Name", "Opt", "ML", "Pess", "Deps"])
    ws.append(["T001", "Design", 3.0, 5.0, 8.0, ""])
    ws.append(["T002", "Development", 10.0, 15.0, 25.0, "T001"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_with_special_characters() -> bytes:
    """Create Excel file with special characters in task names."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    ws.append(["T001", "Design: Phase 1 & 2", 3.0, 5.0, 8.0, ""])
    ws.append(["T002", "Dev (Core) - v1.0", 10.0, 15.0, 25.0, "T001"])
    ws.append(["T003", "Testing @ QA Lab", 5.0, 7.0, 10.0, "T002"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_large_excel(task_count: int = 100) -> bytes:
    """Create Excel file with many tasks."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )

    for i in range(1, task_count + 1):
        task_id = f"T{i:03d}"
        dependencies = f"T{i-1:03d}" if i > 1 else ""
        ws.append([task_id, f"Task {i}", 1.0, 2.0, 3.0, dependencies])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_single_task() -> bytes:
    """Create Excel file with a single task."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    ws.append(["T001", "Single Task", 1.0, 2.0, 3.0, ""])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_no_dependencies() -> bytes:
    """Create Excel file with multiple tasks but no dependencies."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    ws.append(["T001", "Task 1", 1.0, 2.0, 3.0, ""])
    ws.append(["T002", "Task 2", 1.0, 2.0, 3.0, ""])
    ws.append(["T003", "Task 3", 1.0, 2.0, 3.0, ""])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_all_dependent() -> bytes:
    """Create Excel file where all tasks depend on previous task."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"

    ws.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    ws.append(["T001", "Task 1", 1.0, 2.0, 3.0, ""])
    ws.append(["T002", "Task 2", 1.0, 2.0, 3.0, "T001"])
    ws.append(["T003", "Task 3", 1.0, 2.0, 3.0, "T002"])
    ws.append(["T004", "Task 4", 1.0, 2.0, 3.0, "T003"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def create_excel_with_multiple_sheets() -> bytes:
    """Create Excel file with multiple sheets."""
    wb = Workbook()

    # First sheet (not Task List)
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.append(["Project Name", "SprintForge"])

    # Second sheet (Task List)
    ws2 = wb.create_sheet("Task List")
    ws2.append(
        [
            "Task ID",
            "Task Name",
            "Optimistic",
            "Most Likely",
            "Pessimistic",
            "Dependencies",
        ]
    )
    ws2.append(["T001", "Design", 3.0, 5.0, 8.0, ""])
    ws2.append(["T002", "Development", 10.0, 15.0, 25.0, "T001"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# Test Class
class TestExcelParserService:
    """Test suite for ExcelParserService."""

    @pytest.fixture
    def parser_service(self) -> ExcelParserService:
        """Create ExcelParserService instance."""
        return ExcelParserService()

    # Test: Valid Excel File Parsing
    def test_parse_valid_excel(self, parser_service: ExcelParserService):
        """Test parsing a valid Excel file."""
        excel_bytes = create_valid_excel()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")

        assert isinstance(result, ParsedExcelData)
        assert len(result.tasks) == 4
        assert result.tasks[0].task_id == "T001"
        assert result.tasks[0].task_name == "Design"
        assert result.tasks[0].optimistic == 3.0
        assert result.tasks[0].most_likely == 5.0
        assert result.tasks[0].pessimistic == 8.0
        assert result.tasks[0].dependencies == ""
        assert result.tasks[1].dependencies == "T001"

    # Test: Empty File
    def test_parse_empty_file(self, parser_service: ExcelParserService):
        """Test parsing an empty file."""
        with pytest.raises(ExcelParseError, match="Empty file"):
            parser_service.parse_excel_file(b"", "test.xlsx")

    # Test: Invalid Excel Format
    def test_parse_invalid_excel_format(self, parser_service: ExcelParserService):
        """Test parsing a file that's not a valid Excel file."""
        invalid_bytes = b"This is not an Excel file"

        with pytest.raises(ExcelParseError, match="Not a valid Excel file"):
            parser_service.parse_excel_file(invalid_bytes, "test.xlsx")

    # Test: Missing Required Columns
    def test_parse_missing_columns(self, parser_service: ExcelParserService):
        """Test parsing Excel file with missing required columns."""
        excel_bytes = create_excel_with_missing_columns()

        with pytest.raises(ExcelParseError, match="Missing required columns"):
            parser_service.parse_excel_file(excel_bytes, "test.xlsx")

    # Test: Invalid Data Types
    def test_parse_invalid_data_types(self, parser_service: ExcelParserService):
        """Test parsing Excel file with invalid data types."""
        excel_bytes = create_excel_with_invalid_data_types()

        with pytest.raises(ExcelParseError, match="Invalid data type"):
            parser_service.parse_excel_file(excel_bytes, "test.xlsx")

    # Test: PERT Order Violation
    def test_parse_pert_violation(self, parser_service: ExcelParserService):
        """Test parsing Excel file with PERT order violation."""
        excel_bytes = create_excel_with_pert_violation()

        with pytest.raises(ExcelParseError, match="Invalid PERT order"):
            parser_service.parse_excel_file(excel_bytes, "test.xlsx")

    # Test: Circular Dependencies
    def test_parse_circular_dependencies(self, parser_service: ExcelParserService):
        """Test detecting circular dependencies in task data."""
        excel_bytes = create_excel_with_circular_dependencies()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")
        errors = parser_service.validate_task_structure(result.tasks)

        assert len(errors) > 0
        assert any("circular" in error.lower() for error in errors)

    # Test: Empty Task List
    def test_parse_empty_task_list(self, parser_service: ExcelParserService):
        """Test parsing Excel file with headers but no data."""
        excel_bytes = create_empty_excel()

        with pytest.raises(ExcelParseError, match="No tasks found"):
            parser_service.parse_excel_file(excel_bytes, "test.xlsx")

    # Test: Alternative Headers
    def test_parse_alternative_headers(self, parser_service: ExcelParserService):
        """Test parsing Excel file with alternative header names."""
        excel_bytes = create_excel_with_alternative_headers()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")

        assert len(result.tasks) == 2
        assert result.tasks[0].task_id == "T001"
        assert result.tasks[1].dependencies == "T001"

    # Test: Special Characters
    def test_parse_special_characters(self, parser_service: ExcelParserService):
        """Test parsing Excel file with special characters in task names."""
        excel_bytes = create_excel_with_special_characters()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")

        assert len(result.tasks) == 3
        assert ":" in result.tasks[0].task_name
        assert "&" in result.tasks[0].task_name
        assert "(" in result.tasks[1].task_name
        assert "@" in result.tasks[2].task_name

    # Test: Large File
    def test_parse_large_file(self, parser_service: ExcelParserService):
        """Test parsing Excel file with 100+ tasks."""
        excel_bytes = create_large_excel(100)

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")

        assert len(result.tasks) == 100
        assert result.tasks[0].task_id == "T001"
        assert result.tasks[99].task_id == "T100"
        assert result.tasks[99].dependencies == "T099"

    # Test: Single Task
    def test_parse_single_task(self, parser_service: ExcelParserService):
        """Test parsing Excel file with a single task."""
        excel_bytes = create_excel_single_task()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")

        assert len(result.tasks) == 1
        assert result.tasks[0].task_id == "T001"
        assert result.tasks[0].dependencies == ""

    # Test: No Dependencies
    def test_parse_no_dependencies(self, parser_service: ExcelParserService):
        """Test parsing Excel file with no dependencies."""
        excel_bytes = create_excel_no_dependencies()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")

        assert len(result.tasks) == 3
        for task in result.tasks:
            assert task.dependencies == ""

    # Test: All Tasks Dependent
    def test_parse_all_dependent(self, parser_service: ExcelParserService):
        """Test parsing Excel file where all tasks depend on previous task."""
        excel_bytes = create_excel_all_dependent()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")

        assert len(result.tasks) == 4
        assert result.tasks[0].dependencies == ""
        assert result.tasks[1].dependencies == "T001"
        assert result.tasks[2].dependencies == "T002"
        assert result.tasks[3].dependencies == "T003"

    # Test: Multiple Sheets
    def test_parse_multiple_sheets(self, parser_service: ExcelParserService):
        """Test parsing Excel file with multiple sheets (should use 'Task List' sheet)."""
        excel_bytes = create_excel_with_multiple_sheets()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")

        assert len(result.tasks) == 2
        assert result.tasks[0].task_id == "T001"

    # Test: Validate Task Structure - Valid
    def test_validate_task_structure_valid(self, parser_service: ExcelParserService):
        """Test validating a valid task structure."""
        tasks = [
            ParsedTask(
                task_id="T001",
                task_name="Task 1",
                optimistic=1.0,
                most_likely=2.0,
                pessimistic=3.0,
            ),
            ParsedTask(
                task_id="T002",
                task_name="Task 2",
                optimistic=1.0,
                most_likely=2.0,
                pessimistic=3.0,
                dependencies="T001",
            ),
        ]

        errors = parser_service.validate_task_structure(tasks)

        assert len(errors) == 0

    # Test: Validate Task Structure - Circular Dependencies
    def test_validate_task_structure_circular(self, parser_service: ExcelParserService):
        """Test detecting circular dependencies during validation."""
        tasks = [
            ParsedTask(
                task_id="T001",
                task_name="Task 1",
                optimistic=1.0,
                most_likely=2.0,
                pessimistic=3.0,
                dependencies="T003",
            ),
            ParsedTask(
                task_id="T002",
                task_name="Task 2",
                optimistic=1.0,
                most_likely=2.0,
                pessimistic=3.0,
                dependencies="T001",
            ),
            ParsedTask(
                task_id="T003",
                task_name="Task 3",
                optimistic=1.0,
                most_likely=2.0,
                pessimistic=3.0,
                dependencies="T002",
            ),
        ]

        errors = parser_service.validate_task_structure(tasks)

        assert len(errors) > 0
        assert any("circular" in error.lower() for error in errors)

    # Test: Validate Task Structure - Unknown Dependencies
    def test_validate_task_structure_unknown_dependencies(
        self, parser_service: ExcelParserService
    ):
        """Test detecting unknown dependencies during validation."""
        tasks = [
            ParsedTask(
                task_id="T001",
                task_name="Task 1",
                optimistic=1.0,
                most_likely=2.0,
                pessimistic=3.0,
            ),
            ParsedTask(
                task_id="T002",
                task_name="Task 2",
                optimistic=1.0,
                most_likely=2.0,
                pessimistic=3.0,
                dependencies="T999",
            ),
        ]

        errors = parser_service.validate_task_structure(tasks)

        assert len(errors) > 0
        assert any(
            "unknown" in error.lower() or "does not exist" in error.lower()
            for error in errors
        )

    # Test: Convert to Distribution Input
    def test_convert_to_distribution_input(self, parser_service: ExcelParserService):
        """Test converting parsed tasks to TaskDistributionInput."""
        tasks = [
            ParsedTask(
                task_id="T001",
                task_name="Task 1",
                optimistic=1.0,
                most_likely=2.0,
                pessimistic=3.0,
            ),
            ParsedTask(
                task_id="T002",
                task_name="Task 2",
                optimistic=5.0,
                most_likely=7.0,
                pessimistic=10.0,
                dependencies="T001",
            ),
        ]

        distribution_inputs = parser_service.convert_to_distribution_input(tasks)

        assert len(distribution_inputs) == 2
        assert isinstance(distribution_inputs[0], TaskDistributionInput)
        assert distribution_inputs[0].task_id == "T001"
        assert distribution_inputs[0].dependencies == ""
        assert distribution_inputs[1].task_id == "T002"
        assert distribution_inputs[1].dependencies == "T001"

    # Test: ParsedTask PERT Validation
    def test_parsed_task_pert_validation(self):
        """Test PERT order validation in ParsedTask model."""
        # Valid PERT order
        task = ParsedTask(
            task_id="T001",
            task_name="Task 1",
            optimistic=1.0,
            most_likely=2.0,
            pessimistic=3.0,
        )
        assert task.optimistic == 1.0

        # Invalid PERT order
        with pytest.raises(ValueError, match="Invalid PERT order"):
            ParsedTask(
                task_id="T002",
                task_name="Task 2",
                optimistic=3.0,
                most_likely=2.0,
                pessimistic=1.0,
            )

    # Test: Workbook with no active sheet
    def test_parse_workbook_no_sheets(self, parser_service: ExcelParserService):
        """Test handling workbook with no accessible sheets."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"

        # Add headers but wrong sheet
        ws.append(["Not", "Task", "Headers"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        with pytest.raises(ExcelParseError):
            parser_service.parse_excel_file(buffer.read(), "test.xlsx")

    # Test: Multiple dependencies
    def test_parse_multiple_dependencies(self, parser_service: ExcelParserService):
        """Test parsing tasks with multiple dependencies."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Task List"

        ws.append(
            [
                "Task ID",
                "Task Name",
                "Optimistic",
                "Most Likely",
                "Pessimistic",
                "Dependencies",
            ]
        )
        ws.append(["T001", "Task 1", 1.0, 2.0, 3.0, ""])
        ws.append(["T002", "Task 2", 1.0, 2.0, 3.0, ""])
        ws.append(
            ["T003", "Task 3", 1.0, 2.0, 3.0, "T001,T002"]
        )  # Multiple dependencies

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        result = parser_service.parse_excel_file(buffer.read(), "test.xlsx")

        assert len(result.tasks) == 3
        assert result.tasks[2].dependencies == "T001,T002"

        # Validate structure should pass
        errors = parser_service.validate_task_structure(result.tasks)
        assert len(errors) == 0

    # Test: Parse with notes column
    def test_parse_with_notes(self, parser_service: ExcelParserService):
        """Test parsing tasks with notes column."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Task List"

        ws.append(
            [
                "Task ID",
                "Task Name",
                "Optimistic",
                "Most Likely",
                "Pessimistic",
                "Dependencies",
                "Notes",
            ]
        )
        ws.append(["T001", "Task 1", 1.0, 2.0, 3.0, "", "Important task"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        result = parser_service.parse_excel_file(buffer.read(), "test.xlsx")

        assert len(result.tasks) == 1
        assert result.tasks[0].notes == "Important task"


# Integration Tests (if D1 is available)
class TestExcelParserIntegration:
    """Integration tests with Excel generation service (D1)."""

    @pytest.fixture
    def parser_service(self) -> ExcelParserService:
        """Create ExcelParserService instance."""
        return ExcelParserService()

    def test_round_trip_parsing(self, parser_service: ExcelParserService):
        """Test round-trip: Generate Excel → Parse → Validate."""
        # This test will be implemented when D1 is available
        # For now, we'll use our helper function
        excel_bytes = create_valid_excel()

        result = parser_service.parse_excel_file(excel_bytes, "test.xlsx")
        errors = parser_service.validate_task_structure(result.tasks)

        assert len(errors) == 0
        assert len(result.tasks) == 4

        # Convert to distribution input
        distribution_inputs = parser_service.convert_to_distribution_input(result.tasks)
        assert len(distribution_inputs) == 4
