"""
Manual validation test to generate Excel file for human inspection.

Run with: pytest tests/services/test_excel_validation.py -v
"""

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.excel_generation_service import ExcelGenerationService
from app.services.simulation_service import SimulationResult


@pytest.fixture
def output_path():
    """Get output path for validation file."""
    return Path(__file__).parent.parent.parent / "test_monte_carlo_output.xlsx"


def test_generate_validation_excel_file(output_path):
    """
    Generate complete Excel file for manual validation.

    This test creates a full-featured Excel file that can be opened
    in Excel or LibreOffice to verify:
    - All sheets render correctly
    - PERT formulas calculate properly
    - Formatting displays as expected
    - No errors or warnings on open
    """
    # Create service
    service = ExcelGenerationService()

    # Create workbook with all features
    workbook = service.create_template_workbook(
        project_name="Monte Carlo Validation Test", include_sample_data=True
    )

    # Add Monte Carlo Results
    sample_result = SimulationResult(
        project_duration_days=25.3,
        confidence_intervals={10: 18.2, 50: 24.5, 90: 33.1, 95: 36.8, 99: 42.5},
        mean_duration=25.3,
        median_duration=24.5,
        std_deviation=5.2,
        iterations_run=10000,
        simulation_date=datetime.now(),
        task_count=15,
    )

    sample_tasks = [
        {
            "task_id": f"TASK-{i}",
            "task_name": f"Sample Task {i}",
            "optimistic": 2.0 + i * 0.5,
            "most_likely": 4.0 + i * 0.5,
            "pessimistic": 6.0 + i * 0.5,
            "dependencies": f"TASK-{i-1}" if i > 1 and i % 3 == 0 else "",
        }
        for i in range(1, 16)
    ]

    critical_path = ["TASK-1", "TASK-3", "TASK-6", "TASK-12", "TASK-15"]

    service.add_monte_carlo_results_sheet(
        workbook, sample_result, sample_tasks, critical_path
    )

    # Add Quick Simulation sheet
    service.create_quick_simulation_sheet(workbook)

    # Apply formatting
    service.apply_formatting(workbook)

    # Save to file
    excel_bytes = service.save_workbook_to_bytes(workbook)
    with open(output_path, "wb") as f:
        f.write(excel_bytes)

    file_size_kb = len(excel_bytes) / 1024

    # Validate by reopening
    loaded_wb = load_workbook(output_path)

    # Print validation results
    print("\n" + "=" * 70)
    print("EXCEL FILE VALIDATION RESULTS")
    print("=" * 70)
    print(f"\n✓ File created: {output_path}")
    print(f"✓ File size: {file_size_kb:.1f} KB")
    print(f"✓ Sheets: {', '.join(loaded_wb.sheetnames)}")
    print("\nSheet details:")
    for sheet_name in loaded_wb.sheetnames:
        sheet = loaded_wb[sheet_name]
        print(f"  - {sheet_name}: {sheet.max_row} rows × {sheet.max_column} columns")

    # Validate PERT formulas
    task_sheet = loaded_wb["Task List"]
    pert_cell = task_sheet["F2"]
    print(f"\n✓ PERT formula validation:")
    if isinstance(pert_cell.value, str) and "=" in pert_cell.value:
        print(f"    Formula: {pert_cell.value}")
    else:
        print(f"    Cell value: {pert_cell.value}")

    # Validate Monte Carlo Results data
    results_sheet = loaded_wb["Monte Carlo Results"]
    print(f"\n✓ Monte Carlo Results validation:")
    print(f"    Iterations: {results_sheet['B2'].value}")
    print(f"    Mean Duration: {results_sheet['C2'].value} days")
    print(f"    P50: {results_sheet['G2'].value} days")
    print(f"    P90: {results_sheet['H2'].value} days")

    # Validate Quick Simulation
    quick_sim_sheet = loaded_wb["Quick Simulation"]
    print(f"\n✓ Quick Simulation validation:")
    print(f"    Total tasks: {quick_sim_sheet.max_row - 1}")

    print("\n" + "=" * 70)
    print("MANUAL VALIDATION INSTRUCTIONS")
    print("=" * 70)
    print("\n1. Open file in Excel or LibreOffice:")
    print(f"   {output_path}")
    print("\n2. Verify the following:")
    print("   □ File opens without errors or security warnings")
    print("   □ All 3 sheets are present (Task List, Monte Carlo Results, Quick Simulation)")
    print("   □ PERT Mean column (F) shows calculated values")
    print("   □ Headers are bold with light blue background")
    print("   □ PERT Mean column has light green background")
    print("   □ All cells have borders")
    print("   □ Numbers show 2 decimal places")
    print("   □ Header row is frozen (scrolling keeps it visible)")
    print("   □ Monte Carlo Results shows simulation statistics")
    print("   □ Quick Simulation has 100 sample tasks")
    print("   □ Dependencies are properly formatted (comma-separated)")
    print("\n" + "=" * 70)

    # Assertions for automated validation
    assert loaded_wb is not None
    assert len(loaded_wb.sheetnames) == 3
    assert "Task List" in loaded_wb.sheetnames
    assert "Monte Carlo Results" in loaded_wb.sheetnames
    assert "Quick Simulation" in loaded_wb.sheetnames
    assert file_size_kb > 10  # Reasonable file size
    assert quick_sim_sheet.max_row == 101  # Header + 100 tasks
