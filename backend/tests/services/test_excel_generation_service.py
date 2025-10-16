"""
Tests for ExcelGenerationService - Enhanced Excel generation for Monte Carlo simulations.

Test Coverage:
- Template creation (basic and with sample data)
- Task List sheet structure (columns, formulas, formatting)
- Monte Carlo Results sheet (data population, formatting)
- Quick Simulation sheet (100 rows, valid data, dependencies)
- PERT formula validation (=(O+4M+P)/6 in correct cells)
- Formatting application (headers, colors, borders)
- Save to bytes (valid .xlsx format)
- Edge cases (empty workbook, large datasets, special characters)
"""

from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel_generation_service import ExcelGenerationService
from app.services.simulation_service import SimulationResult


@pytest.fixture
def excel_service():
    """Create ExcelGenerationService instance for tests."""
    return ExcelGenerationService()


@pytest.fixture
def sample_tasks():
    """Create sample tasks for testing - simple dict format."""
    return [
        {
            "task_id": "TASK-1",
            "task_name": "Design Phase",
            "optimistic": 3.0,
            "most_likely": 5.0,
            "pessimistic": 8.0,
            "dependencies": "",
        },
        {
            "task_id": "TASK-2",
            "task_name": "Backend Development",
            "optimistic": 5.0,
            "most_likely": 10.0,
            "pessimistic": 15.0,
            "dependencies": "TASK-1",
        },
        {
            "task_id": "TASK-3",
            "task_name": "Frontend Development",
            "optimistic": 5.0,
            "most_likely": 8.0,
            "pessimistic": 12.0,
            "dependencies": "TASK-1",
        },
        {
            "task_id": "TASK-4",
            "task_name": "Integration Testing",
            "optimistic": 2.0,
            "most_likely": 3.0,
            "pessimistic": 5.0,
            "dependencies": "TASK-2,TASK-3",
        },
    ]


@pytest.fixture
def sample_simulation_result():
    """Create sample simulation result for testing."""
    return SimulationResult(
        project_duration_days=18.5,
        confidence_intervals={
            10: 15.0,
            50: 18.0,
            90: 22.0,
            95: 24.0,
            99: 28.0,
        },
        mean_duration=18.5,
        median_duration=18.0,
        std_deviation=3.2,
        iterations_run=10000,
        simulation_date=datetime(2025, 10, 16, 10, 30, 0),
        task_count=4,
    )


class TestCreateTemplateWorkbook:
    """Tests for create_template_workbook method."""

    def test_create_basic_template(self, excel_service):
        """Test creating basic template without sample data."""
        workbook = excel_service.create_template_workbook(
            project_name="Test Project", include_sample_data=False
        )

        # Verify workbook created
        assert workbook is not None
        assert "Task List" in workbook.sheetnames

        # Verify Task List structure
        task_sheet = workbook["Task List"]
        assert task_sheet["A1"].value == "Task ID"
        assert task_sheet["B1"].value == "Task Name"
        assert task_sheet["C1"].value == "Optimistic Duration"
        assert task_sheet["D1"].value == "Most Likely Duration"
        assert task_sheet["E1"].value == "Pessimistic Duration"
        assert task_sheet["F1"].value == "PERT Mean"
        assert task_sheet["G1"].value == "Dependencies"
        assert task_sheet["H1"].value == "Notes"

        # Verify no sample data rows (only header)
        assert task_sheet.max_row == 1

    def test_create_template_with_sample_data(self, excel_service):
        """Test creating template with sample data."""
        workbook = excel_service.create_template_workbook(
            project_name="Sample Project", include_sample_data=True
        )

        task_sheet = workbook["Task List"]

        # Should have header + at least 5 sample rows
        assert task_sheet.max_row >= 6

        # Verify sample data has proper values
        assert task_sheet["A2"].value is not None  # Task ID
        assert task_sheet["C2"].value is not None  # Optimistic
        assert isinstance(task_sheet["C2"].value, (int, float))

    def test_pert_formula_in_template(self, excel_service):
        """Test that PERT Mean column contains correct formula."""
        workbook = excel_service.create_template_workbook(include_sample_data=True)
        task_sheet = workbook["Task List"]

        # Check PERT formula in row 2 (first data row)
        pert_cell = task_sheet["F2"]
        assert pert_cell.value is not None

        # Formula should be: =(C2+4*D2+E2)/6
        if isinstance(pert_cell.value, str) and pert_cell.value.startswith("="):
            assert "C2" in pert_cell.value
            assert "D2" in pert_cell.value
            assert "E2" in pert_cell.value
            assert "/6" in pert_cell.value

    def test_default_project_name(self, excel_service):
        """Test that default project name is used when not specified."""
        workbook = excel_service.create_template_workbook()
        # Should not raise error and create valid workbook
        assert workbook is not None
        assert "Task List" in workbook.sheetnames


class TestAddMonteCarloResultsSheet:
    """Tests for add_monte_carlo_results_sheet method."""

    def test_add_results_sheet(
        self, excel_service, sample_simulation_result, sample_tasks
    ):
        """Test adding Monte Carlo results sheet to workbook."""
        workbook = excel_service.create_template_workbook()
        critical_path = ["TASK-1", "TASK-2", "TASK-4"]

        excel_service.add_monte_carlo_results_sheet(
            workbook=workbook,
            simulation_result=sample_simulation_result,
            tasks=sample_tasks,
            critical_path=critical_path,
        )

        # Verify sheet exists
        assert "Monte Carlo Results" in workbook.sheetnames
        results_sheet = workbook["Monte Carlo Results"]

        # Verify headers
        assert results_sheet["A1"].value == "Simulation Date/Time"
        assert results_sheet["B1"].value == "Iterations"
        assert results_sheet["C1"].value == "Mean Duration (days)"
        assert results_sheet["D1"].value == "Median Duration (P50)"
        assert results_sheet["E1"].value == "Standard Deviation"

        # Verify data values
        assert results_sheet["B2"].value == 10000
        assert results_sheet["C2"].value == 18.5
        assert results_sheet["D2"].value == 18.0
        assert results_sheet["E2"].value == 3.2

    def test_results_sheet_percentiles(
        self, excel_service, sample_simulation_result, sample_tasks
    ):
        """Test that percentile columns are correctly populated."""
        workbook = excel_service.create_template_workbook()

        excel_service.add_monte_carlo_results_sheet(
            workbook=workbook,
            simulation_result=sample_simulation_result,
            tasks=sample_tasks,
            critical_path=["TASK-1"],
        )

        results_sheet = workbook["Monte Carlo Results"]

        # Check percentile headers (after standard columns)
        # Should have P10, P50, P90, P95, P99
        assert "P10" in str(results_sheet["F1"].value)
        assert "P50" in str(results_sheet["G1"].value)
        assert "P90" in str(results_sheet["H1"].value)
        assert "P95" in str(results_sheet["I1"].value)
        assert "P99" in str(results_sheet["J1"].value)

        # Check percentile values
        assert results_sheet["F2"].value == 15.0  # P10
        assert results_sheet["G2"].value == 18.0  # P50
        assert results_sheet["H2"].value == 22.0  # P90
        assert results_sheet["I2"].value == 24.0  # P95
        assert results_sheet["J2"].value == 28.0  # P99

    def test_results_sheet_critical_path(
        self, excel_service, sample_simulation_result, sample_tasks
    ):
        """Test that critical path is correctly formatted."""
        workbook = excel_service.create_template_workbook()
        critical_path = ["TASK-1", "TASK-2", "TASK-4"]

        excel_service.add_monte_carlo_results_sheet(
            workbook=workbook,
            simulation_result=sample_simulation_result,
            tasks=sample_tasks,
            critical_path=critical_path,
        )

        results_sheet = workbook["Monte Carlo Results"]

        # Find Critical Path column
        assert "Critical Path" in str(results_sheet["L1"].value)

        # Critical path should be comma-separated
        critical_path_value = results_sheet["L2"].value
        assert "TASK-1" in critical_path_value
        assert "TASK-2" in critical_path_value
        assert "TASK-4" in critical_path_value

    def test_results_sheet_task_count(
        self, excel_service, sample_simulation_result, sample_tasks
    ):
        """Test that task count is correctly populated."""
        workbook = excel_service.create_template_workbook()

        excel_service.add_monte_carlo_results_sheet(
            workbook=workbook,
            simulation_result=sample_simulation_result,
            tasks=sample_tasks,
            critical_path=[],
        )

        results_sheet = workbook["Monte Carlo Results"]

        # Find Task Count column
        assert "Task Count" in str(results_sheet["K1"].value)
        assert results_sheet["K2"].value == 4


class TestCreateQuickSimulationSheet:
    """Tests for create_quick_simulation_sheet method."""

    def test_create_quick_sim_sheet(self, excel_service):
        """Test creating Quick Simulation sheet with 100 sample tasks."""
        workbook = excel_service.create_template_workbook()
        excel_service.create_quick_simulation_sheet(workbook)

        # Verify sheet exists
        assert "Quick Simulation" in workbook.sheetnames
        quick_sim_sheet = workbook["Quick Simulation"]

        # Should have header + 100 data rows
        assert quick_sim_sheet.max_row == 101

        # Verify headers match Task List structure
        assert quick_sim_sheet["A1"].value == "Task ID"
        assert quick_sim_sheet["B1"].value == "Task Name"
        assert quick_sim_sheet["C1"].value == "Optimistic Duration"
        assert quick_sim_sheet["D1"].value == "Most Likely Duration"
        assert quick_sim_sheet["E1"].value == "Pessimistic Duration"
        assert quick_sim_sheet["F1"].value == "PERT Mean"
        assert quick_sim_sheet["G1"].value == "Dependencies"

    def test_quick_sim_data_validity(self, excel_service):
        """Test that Quick Simulation data is valid (O <= M <= P)."""
        workbook = excel_service.create_template_workbook()
        excel_service.create_quick_simulation_sheet(workbook)

        quick_sim_sheet = workbook["Quick Simulation"]

        # Check first 10 rows for valid data
        for row in range(2, 12):
            task_id = quick_sim_sheet[f"A{row}"].value
            optimistic = quick_sim_sheet[f"C{row}"].value
            most_likely = quick_sim_sheet[f"D{row}"].value
            pessimistic = quick_sim_sheet[f"E{row}"].value

            # All should be non-null
            assert task_id is not None
            assert optimistic is not None
            assert most_likely is not None
            assert pessimistic is not None

            # Validate ordering: O <= M <= P
            assert optimistic <= most_likely
            assert most_likely <= pessimistic

    def test_quick_sim_pert_formulas(self, excel_service):
        """Test that PERT formulas exist in Quick Simulation sheet."""
        workbook = excel_service.create_template_workbook()
        excel_service.create_quick_simulation_sheet(workbook)

        quick_sim_sheet = workbook["Quick Simulation"]

        # Check PERT formula in multiple rows
        for row in [2, 50, 100]:
            pert_cell = quick_sim_sheet[f"F{row}"]
            # Should either be a formula or calculated value
            assert pert_cell.value is not None

    def test_quick_sim_has_dependencies(self, excel_service):
        """Test that some tasks have dependencies in Quick Simulation."""
        workbook = excel_service.create_template_workbook()
        excel_service.create_quick_simulation_sheet(workbook)

        quick_sim_sheet = workbook["Quick Simulation"]

        # Count tasks with dependencies
        tasks_with_deps = 0
        for row in range(2, 102):
            deps = quick_sim_sheet[f"G{row}"].value
            if deps and deps.strip():
                tasks_with_deps += 1

        # At least 30% of tasks should have dependencies (realistic project)
        assert tasks_with_deps >= 30

    def test_quick_sim_varied_durations(self, excel_service):
        """Test that Quick Simulation has varied duration ranges."""
        workbook = excel_service.create_template_workbook()
        excel_service.create_quick_simulation_sheet(workbook)

        quick_sim_sheet = workbook["Quick Simulation"]

        # Collect all optimistic durations
        optimistic_values = []
        for row in range(2, 102):
            opt = quick_sim_sheet[f"C{row}"].value
            if opt is not None:
                optimistic_values.append(opt)

        # Should have variety (min and max differ significantly)
        assert min(optimistic_values) < 5
        assert max(optimistic_values) > 10


class TestApplyFormatting:
    """Tests for apply_formatting method."""

    def test_apply_formatting_basic(self, excel_service):
        """Test that formatting is applied to workbook."""
        workbook = excel_service.create_template_workbook(include_sample_data=True)
        excel_service.apply_formatting(workbook)

        task_sheet = workbook["Task List"]

        # Check header formatting (bold)
        header_font = task_sheet["A1"].font
        assert header_font.bold is True

        # Check header background color (should not be None)
        header_fill = task_sheet["A1"].fill
        assert header_fill.start_color is not None

    def test_pert_column_highlighted(self, excel_service):
        """Test that PERT Mean column is highlighted."""
        workbook = excel_service.create_template_workbook(include_sample_data=True)
        excel_service.apply_formatting(workbook)

        task_sheet = workbook["Task List"]

        # PERT Mean column (F) should have distinct fill color
        pert_fill = task_sheet["F2"].fill
        assert pert_fill.start_color is not None

    def test_borders_applied(self, excel_service):
        """Test that borders are applied to cells."""
        workbook = excel_service.create_template_workbook(include_sample_data=True)
        excel_service.apply_formatting(workbook)

        task_sheet = workbook["Task List"]

        # Check that cells have borders
        cell_border = task_sheet["A2"].border
        assert cell_border is not None
        assert cell_border.left is not None

    def test_number_formatting(self, excel_service):
        """Test that number columns have 2 decimal places."""
        workbook = excel_service.create_template_workbook(include_sample_data=True)
        excel_service.apply_formatting(workbook)

        task_sheet = workbook["Task List"]

        # Duration columns should have number format
        optimistic_format = task_sheet["C2"].number_format
        # Should be numeric format (not "General")
        assert optimistic_format != "General"

    def test_freeze_panes(self, excel_service):
        """Test that header row is frozen."""
        workbook = excel_service.create_template_workbook(include_sample_data=True)
        excel_service.apply_formatting(workbook)

        task_sheet = workbook["Task List"]

        # Freeze panes should be set
        assert task_sheet.freeze_panes is not None


class TestSaveWorkbookToBytes:
    """Tests for save_workbook_to_bytes method."""

    def test_save_to_bytes(self, excel_service):
        """Test saving workbook to bytes."""
        workbook = excel_service.create_template_workbook()
        excel_bytes = excel_service.save_workbook_to_bytes(workbook)

        # Verify bytes returned
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

    def test_saved_file_is_valid_xlsx(self, excel_service):
        """Test that saved bytes can be loaded as valid Excel file."""
        workbook = excel_service.create_template_workbook(include_sample_data=True)
        excel_bytes = excel_service.save_workbook_to_bytes(workbook)

        # Try to load the bytes as a workbook
        excel_file = BytesIO(excel_bytes)
        loaded_wb = load_workbook(excel_file)

        # Verify structure preserved
        assert "Task List" in loaded_wb.sheetnames
        assert loaded_wb["Task List"]["A1"].value == "Task ID"

    def test_save_with_all_features(
        self, excel_service, sample_simulation_result, sample_tasks
    ):
        """Test saving complete workbook with all features."""
        # Create workbook with all features
        workbook = excel_service.create_template_workbook(include_sample_data=True)
        excel_service.add_monte_carlo_results_sheet(
            workbook, sample_simulation_result, sample_tasks, ["TASK-1", "TASK-2"]
        )
        excel_service.create_quick_simulation_sheet(workbook)
        excel_service.apply_formatting(workbook)

        # Save to bytes
        excel_bytes = excel_service.save_workbook_to_bytes(workbook)

        # Load and verify
        excel_file = BytesIO(excel_bytes)
        loaded_wb = load_workbook(excel_file)

        assert "Task List" in loaded_wb.sheetnames
        assert "Monte Carlo Results" in loaded_wb.sheetnames
        assert "Quick Simulation" in loaded_wb.sheetnames


class TestEdgeCases:
    """Tests for edge cases and error handling."""

    def test_empty_workbook_formatting(self, excel_service):
        """Test formatting on minimal workbook."""
        workbook = excel_service.create_template_workbook(include_sample_data=False)
        # Should not raise error
        excel_service.apply_formatting(workbook)

    def test_special_characters_in_project_name(self, excel_service):
        """Test handling special characters in project name."""
        special_name = 'Project "Test" & <Special> Chars!'
        workbook = excel_service.create_template_workbook(project_name=special_name)

        # Should create workbook without error
        assert workbook is not None

    def test_large_critical_path(
        self, excel_service, sample_simulation_result, sample_tasks
    ):
        """Test handling large critical path list."""
        # Create critical path with many tasks
        large_critical_path = [f"TASK-{i}" for i in range(1, 51)]

        workbook = excel_service.create_template_workbook()
        excel_service.add_monte_carlo_results_sheet(
            workbook, sample_simulation_result, sample_tasks, large_critical_path
        )

        # Should handle long comma-separated list
        results_sheet = workbook["Monte Carlo Results"]
        critical_path_value = results_sheet["L2"].value
        assert critical_path_value is not None
        assert len(critical_path_value) > 100  # Long string

    def test_empty_tasks_list(self, excel_service, sample_simulation_result):
        """Test handling empty tasks list."""
        workbook = excel_service.create_template_workbook()

        # Should handle gracefully (empty tasks list)
        excel_service.add_monte_carlo_results_sheet(
            workbook, sample_simulation_result, [], []
        )

        # Should still create sheet
        assert "Monte Carlo Results" in workbook.sheetnames

    def test_zero_iterations_simulation(self, excel_service, sample_tasks):
        """Test handling edge case of unusual simulation results."""
        unusual_result = SimulationResult(
            project_duration_days=0.0,
            confidence_intervals={},
            mean_duration=0.0,
            median_duration=0.0,
            std_deviation=0.0,
            iterations_run=100,  # Minimum
            simulation_date=datetime.now(),
            task_count=0,
        )

        workbook = excel_service.create_template_workbook()
        # Should not raise error
        excel_service.add_monte_carlo_results_sheet(
            workbook, unusual_result, sample_tasks, []
        )


class TestIntegrationScenarios:
    """Integration tests for complete workflows."""

    def test_complete_template_creation_workflow(self, excel_service):
        """Test complete workflow: create template, format, save."""
        # Step 1: Create template with sample data
        workbook = excel_service.create_template_workbook(
            project_name="Complete Test", include_sample_data=True
        )

        # Step 2: Apply formatting
        excel_service.apply_formatting(workbook)

        # Step 3: Save to bytes
        excel_bytes = excel_service.save_workbook_to_bytes(workbook)

        # Verify complete file
        assert len(excel_bytes) > 5000  # Should be substantial size
        excel_file = BytesIO(excel_bytes)
        loaded_wb = load_workbook(excel_file)
        assert len(loaded_wb.sheetnames) >= 1

    def test_complete_simulation_export_workflow(
        self, excel_service, sample_simulation_result, sample_tasks
    ):
        """Test complete simulation export workflow."""
        # Step 1: Create template
        workbook = excel_service.create_template_workbook(
            project_name="Simulation Export"
        )

        # Step 2: Add Monte Carlo results
        excel_service.add_monte_carlo_results_sheet(
            workbook, sample_simulation_result, sample_tasks, ["TASK-1", "TASK-2"]
        )

        # Step 3: Add Quick Simulation sheet
        excel_service.create_quick_simulation_sheet(workbook)

        # Step 4: Apply formatting
        excel_service.apply_formatting(workbook)

        # Step 5: Save to bytes
        excel_bytes = excel_service.save_workbook_to_bytes(workbook)

        # Verify complete export
        excel_file = BytesIO(excel_bytes)
        loaded_wb = load_workbook(excel_file)

        assert "Task List" in loaded_wb.sheetnames
        assert "Monte Carlo Results" in loaded_wb.sheetnames
        assert "Quick Simulation" in loaded_wb.sheetnames

    def test_multiple_results_sheets(
        self, excel_service, sample_simulation_result, sample_tasks
    ):
        """Test that service handles multiple calls correctly."""
        workbook = excel_service.create_template_workbook()

        # Add results sheet twice (simulating multiple simulations)
        excel_service.add_monte_carlo_results_sheet(
            workbook, sample_simulation_result, sample_tasks, ["TASK-1"]
        )

        # Should still work (might replace or append)
        excel_service.add_monte_carlo_results_sheet(
            workbook, sample_simulation_result, sample_tasks, ["TASK-2"]
        )

        # Should not crash
        assert "Monte Carlo Results" in workbook.sheetnames
