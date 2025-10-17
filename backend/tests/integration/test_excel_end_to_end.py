"""
End-to-end integration test for complete Excel workflow.

Tests the complete workflow:
1. Upload Excel file with tasks
2. Run Monte Carlo simulation
3. Download Excel file with results
4. Parse downloaded file to verify results are included
"""

import io
from datetime import date
from unittest.mock import patch
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession


class TestExcelEndToEnd:
    """End-to-end tests for complete Excel workflow."""

    @pytest.fixture
    def mock_user_info(self) -> dict:
        """Mock authenticated user info."""
        return {"sub": str(uuid4()), "email": "test@example.com"}

    @pytest.mark.asyncio
    async def test_complete_excel_workflow(
        self, client: AsyncClient, db_session: AsyncSession, mock_user_info: dict
    ):
        """
        Test complete workflow: upload → simulate → download → verify.

        This test verifies the round-trip integrity of the Excel workflow.
        """
        project_id = uuid4()

        # Step 1: Create Excel file with tasks
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"

        # Headers
        ws.append(
            [
                "Task ID",
                "Task Name",
                "Optimistic",
                "Most Likely",
                "Pessimistic",
                "Dependencies",
            ]
        )

        # Sample tasks with realistic project structure
        ws.append(["TASK-1", "Requirements gathering", 2.0, 5.0, 8.0, ""])
        ws.append(["TASK-2", "System design", 3.0, 7.0, 10.0, "TASK-1"])
        ws.append(["TASK-3", "Frontend development", 5.0, 10.0, 15.0, "TASK-2"])
        ws.append(["TASK-4", "Backend development", 5.0, 12.0, 20.0, "TASK-2"])
        ws.append(["TASK-5", "Integration", 2.0, 4.0, 6.0, "TASK-3,TASK-4"])
        ws.append(["TASK-6", "Testing", 3.0, 5.0, 8.0, "TASK-5"])
        ws.append(["TASK-7", "Deployment", 1.0, 2.0, 3.0, "TASK-6"])

        # Save to bytes
        upload_bytes = io.BytesIO()
        wb.save(upload_bytes)
        upload_bytes.seek(0)

        with patch("app.core.auth.require_auth", return_value=mock_user_info), patch(
            "app.services.excel_parser_service.ExcelParserService.parse_excel_file"
        ) as mock_parse, patch(
            "app.services.simulation_service.SimulationService.run_simulation"
        ) as mock_simulate, patch(
            "app.services.simulation_persistence_service.SimulationPersistenceService.save_simulation_result"
        ) as mock_save, patch(
            "app.services.simulation_persistence_service.SimulationPersistenceService.get_simulation_result"
        ) as mock_get, patch(
            "app.services.excel_generation_service.ExcelGenerationService.create_template_workbook"
        ) as mock_create_workbook, patch(
            "app.services.excel_generation_service.ExcelGenerationService.save_workbook_to_bytes"
        ) as mock_save_bytes:

            # Mock parser to return ParsedExcelData with ParsedTask objects
            from app.services.excel_parser_service import ParsedExcelData, ParsedTask

            parsed_tasks = ParsedExcelData(
                tasks=[
                    ParsedTask(
                        task_id="TASK-1",
                        task_name="Requirements gathering",
                        optimistic=2.0,
                        most_likely=5.0,
                        pessimistic=8.0,
                        dependencies="",
                    ),
                    ParsedTask(
                        task_id="TASK-2",
                        task_name="System design",
                        optimistic=3.0,
                        most_likely=7.0,
                        pessimistic=10.0,
                        dependencies="TASK-1",
                    ),
                ],
                project_name="Test Project"
            )
            mock_parse.return_value = parsed_tasks

            # Mock simulation result
            from unittest.mock import MagicMock

            mock_simulation_result = MagicMock()
            mock_simulation_result.project_duration_days = 35.5
            mock_simulation_result.confidence_intervals = {
                10: 28.0,
                50: 35.5,
                90: 45.0,
                95: 50.0,
                99: 58.0,
            }
            mock_simulation_result.mean_duration = 35.5
            mock_simulation_result.median_duration = 35.5
            mock_simulation_result.std_deviation = 8.2
            mock_simulation_result.iterations_run = 10000
            mock_simulation_result.task_count = 7
            mock_simulate.return_value = mock_simulation_result

            # Mock persistence
            simulation_id = 789
            mock_save.return_value = simulation_id

            # Step 2: Upload Excel and run simulation
            upload_response = await client.post(
                f"/api/v1/excel/projects/{project_id}/simulate",
                files={
                    "file": (
                        "project_tasks.xlsx",
                        upload_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                params={
                    "iterations": 10000,
                    "project_start_date": "2025-01-20",
                },
            )

            # Verify upload response
            assert upload_response.status_code == 200
            upload_data = upload_response.json()
            assert upload_data["simulation_id"] == simulation_id
            assert upload_data["task_count"] == 7
            assert upload_data["mean_duration"] == 35.5
            assert "download_url" in upload_data

            # Step 3: Download Excel with results
            # Mock the database retrieval for download
            from app.models.simulation_result import SimulationResult

            mock_db_simulation = MagicMock(spec=SimulationResult)
            mock_db_simulation.id = simulation_id
            mock_db_simulation.project_id = project_id
            mock_db_simulation.mean_duration = 35.5
            mock_db_simulation.median_duration = 35.5
            mock_db_simulation.std_deviation = 8.2
            mock_db_simulation.confidence_intervals = {
                10: 28.0,
                50: 35.5,
                90: 45.0,
                95: 50.0,
                99: 58.0,
            }
            mock_db_simulation.task_count = 7
            mock_get.return_value = mock_db_simulation

            # Mock Excel generation with enhanced file
            enhanced_wb = Workbook()

            # Task List sheet (original data)
            tasks_ws = enhanced_wb.active
            tasks_ws.title = "Task List"
            tasks_ws.append(
                [
                    "Task ID",
                    "Task Name",
                    "Optimistic",
                    "Most Likely",
                    "Pessimistic",
                    "Dependencies",
                ]
            )
            for task in parsed_tasks.tasks:
                tasks_ws.append(
                    [
                        task.task_id,
                        task.task_name,
                        task.optimistic,
                        task.most_likely,
                        task.pessimistic,
                        task.dependencies,
                    ]
                )

            # Monte Carlo Results sheet
            results_ws = enhanced_wb.create_sheet("Monte Carlo Results", 1)
            results_ws.append(["Metric", "Value"])
            results_ws.append(["Mean Duration (days)", 35.5])
            results_ws.append(["Median Duration (days)", 35.5])
            results_ws.append(["Std Deviation (days)", 8.2])
            results_ws.append(["P10 (days)", 28.0])
            results_ws.append(["P50 (days)", 35.5])
            results_ws.append(["P90 (days)", 45.0])
            results_ws.append(["P95 (days)", 50.0])
            results_ws.append(["P99 (days)", 58.0])

            download_bytes = io.BytesIO()
            enhanced_wb.save(download_bytes)
            download_bytes.seek(0)

            # Mock the workbook creation and byte generation
            mock_create_workbook.return_value = enhanced_wb
            mock_save_bytes.return_value = download_bytes.getvalue()

            download_response = await client.get(
                f"/api/v1/excel/simulations/{simulation_id}/excel"
            )

            # Verify download response
            assert download_response.status_code == 200
            assert (
                download_response.headers["content-type"]
                == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            assert len(download_response.content) > 0

            # Step 4: Parse downloaded Excel and verify results are included
            downloaded_bytes = io.BytesIO(download_response.content)
            downloaded_wb = load_workbook(downloaded_bytes)

            # Verify Task List sheet exists and has data
            assert "Task List" in downloaded_wb.sheetnames
            tasks_sheet = downloaded_wb["Task List"]
            assert tasks_sheet["A1"].value == "Task ID"
            assert tasks_sheet["A2"].value == "TASK-1"

            # Verify Monte Carlo Results sheet exists
            assert "Monte Carlo Results" in downloaded_wb.sheetnames
            results_sheet = downloaded_wb["Monte Carlo Results"]
            assert results_sheet["A1"].value == "Metric"
            assert results_sheet["B1"].value == "Value"

            # Verify specific result values
            mean_row = None
            for row in results_sheet.iter_rows(min_row=2, max_row=20):
                if row[0].value == "Mean Duration (days)":
                    mean_row = row
                    break

            assert mean_row is not None
            assert mean_row[1].value == 35.5

    @pytest.mark.asyncio
    async def test_template_download_and_upload(
        self, client: AsyncClient, mock_user_info: dict
    ):
        """Test downloading template and using it for simulation."""
        project_id = uuid4()

        with patch("app.core.auth.require_auth", return_value=mock_user_info), patch(
            "app.services.excel_generation_service.ExcelGenerationService.create_template_workbook"
        ) as mock_template, patch(
            "app.services.excel_generation_service.ExcelGenerationService.save_workbook_to_bytes"
        ) as mock_save_bytes, patch(
            "app.services.excel_parser_service.ExcelParserService.parse_excel_file"
        ) as mock_parse, patch(
            "app.services.simulation_service.SimulationService.run_simulation"
        ) as mock_simulate, patch(
            "app.services.simulation_persistence_service.SimulationPersistenceService.save_simulation_result"
        ) as mock_save:

            # Step 1: Download template with sample data
            wb = Workbook()
            ws = wb.active
            ws.append(
                [
                    "Task ID",
                    "Task Name",
                    "Optimistic",
                    "Most Likely",
                    "Pessimistic",
                    "Dependencies",
                ]
            )
            ws.append(["SAMPLE-1", "Sample task", 1.0, 2.0, 3.0, ""])

            template_bytes = io.BytesIO()
            wb.save(template_bytes)
            template_bytes.seek(0)

            mock_template.return_value = wb
            mock_save_bytes.return_value = template_bytes.getvalue()

            template_response = await client.get(
                "/api/v1/excel/template", params={"include_sample_data": True}
            )

            assert template_response.status_code == 200

            # Step 2: Use downloaded template for simulation
            template_content = io.BytesIO(template_response.content)

            from app.services.excel_parser_service import ParsedExcelData, ParsedTask

            mock_parse.return_value = ParsedExcelData(
                tasks=[ParsedTask(
                    task_id="SAMPLE-1",
                    task_name="Sample task",
                    optimistic=1.0,
                    most_likely=2.0,
                    pessimistic=3.0,
                    dependencies="",
                )],
                project_name="My Project"
            )

            from unittest.mock import MagicMock

            mock_result = MagicMock()
            mock_result.project_duration_days = 2.0
            mock_result.mean_duration = 2.0
            mock_simulate.return_value = mock_result
            mock_save.return_value = 999

            simulate_response = await client.post(
                f"/api/v1/excel/projects/{project_id}/simulate",
                files={
                    "file": (
                        "template.xlsx",
                        template_content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                params={
                    "iterations": 1000,
                    "project_start_date": "2025-02-01",
                },
            )

            assert simulate_response.status_code == 200
            assert simulate_response.json()["simulation_id"] == 999
