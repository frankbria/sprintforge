"""
Performance Benchmarking Tests for Excel Generation Engine (Task 3.7).

Measures generation time, memory usage, and workbook loading speed
for the Excel template engine.

Note: These tests measure performance of the actual implemented functionality,
which generates a fixed template structure with metadata and sync capabilities.
"""

import pytest
import time
import tracemalloc
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

from app.excel.engine import ExcelTemplateEngine, ProjectConfig


class TestGenerationPerformance:
    """Test Excel template generation time."""

    def test_basic_template_generation_time(self):
        """Test generation time for basic template."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="perf_basic",
            project_name="Basic Performance Test",
        )

        start_time = time.time()
        excel_bytes = engine.generate_template(config)
        generation_time = time.time() - start_time

        assert excel_bytes is not None
        assert generation_time < 1.0  # Should complete in under 1 second
        print(f"\n✓ Basic template generation: {generation_time:.3f}s")

    def test_template_with_features_generation_time(self):
        """Test generation time for template with features."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="perf_features",
            project_name="Features Performance Test",
            features={"gantt_chart": True, "monte_carlo": True, "resource_leveling": True},
        )

        start_time = time.time()
        excel_bytes = engine.generate_template(config)
        generation_time = time.time() - start_time

        assert excel_bytes is not None
        assert generation_time < 1.0  # Should still be fast
        print(f"✓ Template with features: {generation_time:.3f}s")

    def test_generation_consistency(self):
        """Test that generation time is consistent across multiple runs."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="perf_consistency",
            project_name="Consistency Test",
        )

        times = []
        for _ in range(5):
            start_time = time.time()
            engine.generate_template(config)
            times.append(time.time() - start_time)

        avg_time = sum(times) / len(times)
        max_deviation = max(abs(t - avg_time) for t in times)

        # Variance should be reasonable (within 50% of average)
        assert max_deviation < avg_time * 0.5
        print(f"\n✓ Consistency check: avg={avg_time:.3f}s, max_dev={max_deviation:.3f}s")


class TestMemoryUsage:
    """Test memory usage patterns during Excel generation."""

    def test_basic_template_memory_usage(self):
        """Test memory usage for basic template generation."""
        tracemalloc.start()

        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="mem_basic",
            project_name="Basic Memory Test",
        )

        excel_bytes = engine.generate_template(config)

        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()

        # Peak memory should be reasonable (under 20MB for basic template)
        peak_mb = peak / (1024 * 1024)
        assert peak_mb < 20.0
        print(f"\n✓ Basic template memory: {peak_mb:.2f}MB")

    def test_memory_cleanup_after_generation(self):
        """Test that memory is properly released after generation."""
        tracemalloc.start()

        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="mem_cleanup",
            project_name="Memory Cleanup Test",
        )

        # Generate file
        excel_bytes = engine.generate_template(config)

        # Get memory after generation
        after_gen, peak_gen = tracemalloc.get_traced_memory()

        # Clear references
        del excel_bytes
        del engine

        # Force garbage collection
        import gc
        gc.collect()

        # Get memory after cleanup
        after_cleanup, _ = tracemalloc.get_traced_memory()
        tracemalloc.stop()

        # Memory should decrease significantly after cleanup
        cleanup_ratio = after_cleanup / after_gen if after_gen > 0 else 0
        assert cleanup_ratio < 0.7  # At least 30% memory released

        print(f"\n✓ Memory cleanup: {after_gen/(1024*1024):.2f}MB → {after_cleanup/(1024*1024):.2f}MB ({cleanup_ratio*100:.1f}%)")

    def test_concurrent_generation_memory_isolation(self):
        """Test that concurrent generations don't share memory improperly."""
        tracemalloc.start()

        engine1 = ExcelTemplateEngine()
        engine2 = ExcelTemplateEngine()

        config1 = ProjectConfig(
            project_id="mem_concurrent_1",
            project_name="Concurrent Test 1",
        )

        config2 = ProjectConfig(
            project_id="mem_concurrent_2",
            project_name="Concurrent Test 2",
        )

        # Generate both
        bytes1 = engine1.generate_template(config1)
        bytes2 = engine2.generate_template(config2)

        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()

        # Peak should be reasonable for two concurrent generations
        peak_mb = peak / (1024 * 1024)
        assert peak_mb < 40.0  # Less than 2x single generation with overhead

        assert bytes1 != bytes2  # Different outputs
        print(f"\n✓ Concurrent generation memory: {peak_mb:.2f}MB")


class TestWorkbookLoading:
    """Test workbook loading and parsing performance."""

    def test_workbook_loading_time(self):
        """Test time to load generated workbook."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="load_test",
            project_name="Loading Speed Test",
        )

        excel_bytes = engine.generate_template(config)
        buffer = BytesIO(excel_bytes)

        start_time = time.time()
        workbook = load_workbook(buffer)
        load_time = time.time() - start_time

        assert workbook is not None
        assert load_time < 1.0  # Should load quickly
        print(f"\n✓ Workbook loading: {load_time:.3f}s")

    def test_metadata_extraction_performance(self):
        """Test metadata extraction performance."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="metadata_perf",
            project_name="Metadata Performance Test",
        )

        excel_bytes = engine.generate_template(config)

        start_time = time.time()
        metadata = engine.load_metadata_from_excel(excel_bytes)
        extraction_time = time.time() - start_time

        assert metadata is not None
        assert extraction_time < 0.5  # Should be very fast
        print(f"\n✓ Metadata extraction: {extraction_time:.3f}s")

    def test_multiple_load_cycles(self):
        """Test performance of multiple load cycles."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="multi_load",
            project_name="Multi Load Test",
        )

        excel_bytes = engine.generate_template(config)

        load_times = []
        for _ in range(10):
            buffer = BytesIO(excel_bytes)
            start_time = time.time()
            workbook = load_workbook(buffer)
            load_times.append(time.time() - start_time)

        avg_time = sum(load_times) / len(load_times)
        assert avg_time < 1.0  # Average should be fast
        print(f"\n✓ Average load time (10 cycles): {avg_time:.3f}s")


class TestTemplateFeatures:
    """Test performance impact of different template features."""

    def test_minimal_template_performance(self):
        """Test performance of minimal template (no extra features)."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="minimal_perf",
            project_name="Minimal Template",
            features={},
        )

        start_time = time.time()
        excel_bytes = engine.generate_template(config)
        generation_time = time.time() - start_time

        assert generation_time < 0.5  # Minimal should be very fast
        print(f"\n✓ Minimal template: {generation_time:.3f}s")

    def test_full_featured_template_performance(self):
        """Test performance of template with all features enabled."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="full_featured",
            project_name="Full Featured Template",
            features={
                "gantt_chart": True,
                "monte_carlo": True,
                "resource_leveling": True,
                "critical_path": True,
                "earned_value": True,
            },
        )

        start_time = time.time()
        excel_bytes = engine.generate_template(config)
        generation_time = time.time() - start_time

        assert generation_time < 1.5  # Should still be reasonably fast
        print(f"\n✓ Full featured template: {generation_time:.3f}s")

    def test_metadata_overhead(self):
        """Test overhead of metadata generation and embedding."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="metadata_overhead",
            project_name="Metadata Overhead Test",
            metadata={"key1": "value1", "key2": "value2", "key3": "value3"},
        )

        start_time = time.time()
        excel_bytes = engine.generate_template(config)
        total_time = time.time() - start_time

        # Metadata should add minimal overhead
        assert total_time < 1.0
        print(f"\n✓ Template with metadata: {total_time:.3f}s")


class TestPerformanceRegression:
    """Test for performance regressions."""

    def test_performance_baseline(self):
        """Test that performance meets baseline expectations."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="regression_baseline",
            project_name="Regression Baseline",
        )

        times = []
        for _ in range(5):  # Run 5 times
            start_time = time.time()
            engine.generate_template(config)
            times.append(time.time() - start_time)

        avg_time = sum(times) / len(times)
        max_time = max(times)
        min_time = min(times)

        assert avg_time < 0.5  # Average should be very fast
        assert max_time < 1.0  # Even worst case should be under 1 second

        print(f"\n✓ Performance baseline:")
        print(f"  Average: {avg_time:.3f}s")
        print(f"  Max: {max_time:.3f}s")
        print(f"  Min: {min_time:.3f}s")

    def test_file_size_efficiency(self):
        """Test that generated file sizes are reasonable."""
        engine = ExcelTemplateEngine()

        configs = [
            ProjectConfig("size_minimal", "Minimal", features={}),
            ProjectConfig("size_basic", "Basic"),
            ProjectConfig(
                "size_full",
                "Full Featured",
                features={"gantt_chart": True, "monte_carlo": True},
            ),
        ]

        file_sizes = []
        for config in configs:
            excel_bytes = engine.generate_template(config)
            file_size_kb = len(excel_bytes) / 1024
            file_sizes.append((config.project_id, file_size_kb))

        print(f"\n✓ File size efficiency:")
        for project_id, size_kb in file_sizes:
            # Basic template should be under 100KB
            assert size_kb < 100, f"{project_id} too large: {size_kb:.1f}KB"
            print(f"  {project_id}: {size_kb:.1f}KB")

    def test_checksum_calculation_performance(self):
        """Test checksum calculation doesn't add significant overhead."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="checksum_perf",
            project_name="Checksum Performance Test",
        )

        # Generate and time the whole process
        start_time = time.time()
        excel_bytes = engine.generate_template(config)
        total_time = time.time() - start_time

        # Extract metadata to verify checksum exists
        metadata = engine.load_metadata_from_excel(excel_bytes)
        assert "checksum" in metadata

        # Total time should still be fast
        assert total_time < 1.0
        print(f"\n✓ Generation with checksum: {total_time:.3f}s")


class TestEngineScalability:
    """Test engine scalability with multiple projects."""

    def test_sequential_generation_performance(self):
        """Test performance of generating multiple projects sequentially."""
        engine = ExcelTemplateEngine()

        num_projects = 10
        start_time = time.time()

        for i in range(num_projects):
            config = ProjectConfig(
                project_id=f"seq_{i}",
                project_name=f"Sequential Project {i}",
            )
            excel_bytes = engine.generate_template(config)
            assert excel_bytes is not None

        total_time = time.time() - start_time
        avg_time = total_time / num_projects

        assert avg_time < 1.0  # Average time per project should be fast
        print(f"\n✓ Sequential generation ({num_projects} projects):")
        print(f"  Total: {total_time:.3f}s")
        print(f"  Average per project: {avg_time:.3f}s")

    def test_engine_reuse_performance(self):
        """Test that reusing engine instance doesn't degrade performance."""
        engine = ExcelTemplateEngine()

        times = []
        for i in range(10):
            config = ProjectConfig(
                project_id=f"reuse_{i}",
                project_name=f"Reuse Test {i}",
            )
            start_time = time.time()
            engine.generate_template(config)
            times.append(time.time() - start_time)

        # Performance should remain consistent
        first_half_avg = sum(times[:5]) / 5
        second_half_avg = sum(times[5:]) / 5

        # Second half shouldn't be significantly slower (allow 20% variance)
        ratio = second_half_avg / first_half_avg
        assert 0.8 <= ratio <= 1.2

        print(f"\n✓ Engine reuse performance:")
        print(f"  First half avg: {first_half_avg:.3f}s")
        print(f"  Second half avg: {second_half_avg:.3f}s")
        print(f"  Ratio: {ratio:.2f}")


# Performance benchmark summary
def test_performance_benchmark_summary():
    """Summary of all performance benchmarks."""
    # This is a marker test to document performance targets
    assert True
    print("""
    ═══════════════════════════════════════════════════════════════
    Performance Benchmark Targets (Actual Implementation)
    ═══════════════════════════════════════════════════════════════

    Generation Time:
    - Basic template: < 0.5 seconds
    - Full featured template: < 1.5 seconds
    - Average per project: < 1.0 seconds

    Memory Usage:
    - Basic template: < 20 MB peak
    - Concurrent generations: < 40 MB peak
    - Memory cleanup: > 30% released after cleanup

    Workbook Loading:
    - Initial load: < 1.0 seconds
    - Metadata extraction: < 0.5 seconds
    - Average load (repeated): < 1.0 seconds

    File Size:
    - Basic template: < 100 KB
    - Full featured: < 100 KB
    - Scales reasonably with features

    Scalability:
    - Sequential generation: consistent performance
    - Engine reuse: no performance degradation
    - Multiple projects: < 1s average per project

    ═══════════════════════════════════════════════════════════════
    Note: These tests measure the actual implemented functionality,
    which generates fixed template structures with metadata and
    sync capabilities, not dynamic task-based generation.
    ═══════════════════════════════════════════════════════════════
    """)
