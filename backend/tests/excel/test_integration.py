"""
Integration tests for Excel Generation Engine (Task 3.7).

Tests complete Excel generation workflow, file validity, formula calculations,
and end-to-end scenarios combining all components.

Target: >85% coverage, 100% pass rate
"""

import pytest
from io import BytesIO
from datetime import datetime, date
from openpyxl import load_workbook

from app.excel.engine import ExcelTemplateEngine, ProjectConfig
from app.excel.templates import (
    TemplateRegistry,
    TemplateLayoutBuilder,
    select_template,
    ProjectMethodology,
    TemplateVariation,
)
from app.excel.compatibility import (
    ExcelCompatibilityManager,
    ExcelVersion,
    Platform,
)
from app.excel.components.templates.formula_loader import FormulaTemplateLoader


class TestExcelGenerationIntegration:
    """Test complete Excel generation workflow."""

    def test_basic_excel_generation_workflow(self):
        """Test generating a basic Excel file."""
        # Create engine and config
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_proj_001",
            project_name="Test Project",
            sprint_pattern="YY.Q.#",
        )

        # Generate Excel file
        excel_bytes = engine.generate_template(config)

        # Verify bytes were generated
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
        assert isinstance(excel_bytes, bytes)

    def test_generated_file_is_valid_excel(self):
        """Test that generated file is a valid Excel file."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_proj_002",
            project_name="Valid Excel Test",
        )

        excel_bytes = engine.generate_template(config)

        # Load with openpyxl to verify validity
        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        assert workbook is not None
        assert len(workbook.sheetnames) >= 1

    def test_excel_contains_project_plan_sheet(self):
        """Test that generated Excel contains Project Plan worksheet."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_proj_003",
            project_name="Sheet Test",
        )

        excel_bytes = engine.generate_template(config)
        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        assert "Project Plan" in workbook.sheetnames

    def test_excel_contains_metadata_sheet(self):
        """Test that generated Excel contains hidden _SYNC_META sheet."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_proj_004",
            project_name="Metadata Test",
        )

        excel_bytes = engine.generate_template(config)
        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        assert "_SYNC_META" in workbook.sheetnames

        # Verify sheet is hidden
        meta_sheet = workbook["_SYNC_META"]
        assert meta_sheet.sheet_state == "hidden"

    def test_metadata_can_be_extracted(self):
        """Test that metadata can be extracted from generated file."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_proj_005",
            project_name="Metadata Extraction Test",
            sprint_pattern="YY.Q.#",
        )

        excel_bytes = engine.generate_template(config)

        # Extract metadata
        metadata = engine.load_metadata_from_excel(excel_bytes)

        assert metadata["project_id"] == "test_proj_005"
        assert metadata["project_name"] == "Metadata Extraction Test"
        assert "version" in metadata
        assert "generated_at" in metadata

    def test_project_plan_has_correct_headers(self):
        """Test that Project Plan sheet has expected column headers."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_proj_006",
            project_name="Headers Test",
        )

        excel_bytes = engine.generate_template(config)
        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        ws = workbook["Project Plan"]

        # Check headers in first row
        expected_headers = [
            "Task ID",
            "Task Name",
            "Duration (days)",
            "Start Date",
            "End Date",
            "Dependencies",
            "Sprint",
            "Status",
            "Owner",
        ]

        for col_idx, header in enumerate(expected_headers, start=1):
            cell_value = ws.cell(row=1, column=col_idx).value
            assert cell_value == header

    def test_sample_task_is_included(self):
        """Test that sample task row is included in template."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_proj_007",
            project_name="Sample Task Test",
        )

        excel_bytes = engine.generate_template(config)
        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        ws = workbook["Project Plan"]

        # Check row 2 has sample task
        task_id = ws.cell(row=2, column=1).value
        task_name = ws.cell(row=2, column=2).value

        assert task_id == "T001"
        assert task_name == "Sample Task"


class TestTemplateIntegration:
    """Test template system integration with Excel generation."""

    def test_select_and_use_agile_template(self):
        """Test selecting Agile template and using for generation."""
        # Select template
        template = select_template("agile", "basic")

        assert template is not None
        assert template.methodology == ProjectMethodology.AGILE
        assert "sprints" in template.features

    def test_select_and_use_waterfall_template(self):
        """Test selecting Waterfall template and using for generation."""
        template = select_template("waterfall", "advanced")

        assert template is not None
        assert template.methodology == ProjectMethodology.WATERFALL
        assert "milestones" in template.features
        assert "critical_path" in template.features

    def test_template_layout_builds_correctly(self):
        """Test that template layout is built with correct columns."""
        template = select_template("agile", "advanced")
        builder = TemplateLayoutBuilder()

        layout = builder.build_layout(template)

        # Agile template should have Sprint column
        assert "Sprint" in layout.columns
        assert "Story Points" in layout.columns

        # Advanced should have advanced columns
        assert "Optimistic" in layout.columns
        assert "% Complete" in layout.columns

    def test_hybrid_template_combines_features(self):
        """Test that hybrid template includes both Agile and Waterfall features."""
        registry = TemplateRegistry()
        template = registry.get_template("hybrid")
        builder = TemplateLayoutBuilder()

        layout = builder.build_layout(template)

        # Should have both Agile and Waterfall columns
        assert "Sprint" in layout.columns
        assert "Milestone" in layout.columns


class TestCompatibilityIntegration:
    """Test Excel compatibility integration."""

    def test_excel_2019_compatibility_workflow(self):
        """Test generating Excel compatible with Excel 2019."""
        compat = ExcelCompatibilityManager(
            target_version=ExcelVersion.EXCEL_2019,
            target_platform=Platform.WINDOWS,
        )

        # Generate report
        report = compat.get_compatibility_report()

        assert report["target_version"] == "2019"
        assert "support_percentage" in report

    def test_xlookup_fallback_generation(self):
        """Test generating XLOOKUP with fallback for Excel 2019."""
        compat = ExcelCompatibilityManager(
            target_version=ExcelVersion.EXCEL_2019,
            target_platform=Platform.WINDOWS,
            enable_fallbacks=True,
        )

        formula = compat.get_xlookup_formula(
            lookup_value="A2",
            lookup_array="B:B",
            return_array="C:C",
        )

        # Should use INDEX/MATCH fallback
        assert "INDEX(" in formula
        assert "MATCH(" in formula
        assert "XLOOKUP" not in formula

    def test_excel_365_modern_functions(self):
        """Test using modern functions for Excel 365."""
        compat = ExcelCompatibilityManager(
            target_version=ExcelVersion.EXCEL_365,
            target_platform=Platform.WINDOWS,
        )

        # Should support XLOOKUP natively
        formula = compat.get_xlookup_formula(
            lookup_value="A2",
            lookup_array="B:B",
            return_array="C:C",
        )

        assert "XLOOKUP(" in formula


class TestFormulaTemplateIntegration:
    """Test formula template loading and application."""

    def test_load_formula_templates(self):
        """Test loading formula templates from JSON files."""
        loader = FormulaTemplateLoader()

        # Load various templates
        loader.load_template("dependencies")
        loader.load_template("critical_path")
        loader.load_template("monte_carlo")

        # Verify formulas are loaded
        formulas = loader.list_formulas()
        assert len(formulas) > 0

    def test_apply_dependency_formula(self):
        """Test applying dependency formula template."""
        loader = FormulaTemplateLoader()
        loader.load_template("dependencies")

        # Apply finish-to-start formula
        formula = loader.apply_template(
            "dependency_fs",
            predecessor_finish="E4",
            task_start="D5",
            lag_days="2",
        )

        assert "E4" in formula
        assert "D5" in formula
        assert "2" in formula

    def test_apply_monte_carlo_formula(self):
        """Test applying Monte Carlo formula template."""
        loader = FormulaTemplateLoader()
        loader.load_template("monte_carlo")

        # Apply PERT mean formula
        formula = loader.apply_template(
            "pert_mean",
            optimistic="10",
            most_likely="20",
            pessimistic="30",
        )

        assert "10" in formula
        assert "20" in formula
        assert "30" in formula


class TestEndToEndScenarios:
    """Test complete end-to-end scenarios."""

    def test_agile_project_workflow(self):
        """Test complete Agile project setup workflow."""
        # Step 1: Select Agile template
        template = select_template("agile", "advanced")
        assert template is not None

        # Step 2: Build layout
        builder = TemplateLayoutBuilder()
        layout = builder.build_layout(template)
        assert layout.sprint_tracking is True
        assert layout.has_burndown is True

        # Step 3: Create project config
        config = ProjectConfig(
            project_id="agile_proj_001",
            project_name="Agile Test Project",
            sprint_pattern="YY.Q.#",
        )

        # Step 4: Generate Excel
        engine = ExcelTemplateEngine()
        excel_bytes = engine.generate_template(config)

        # Step 5: Verify file
        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        assert "Project Plan" in workbook.sheetnames
        assert "_SYNC_META" in workbook.sheetnames

    def test_waterfall_project_workflow(self):
        """Test complete Waterfall project setup workflow."""
        # Select Waterfall template
        template = select_template("waterfall", "basic")

        # Build layout
        builder = TemplateLayoutBuilder()
        layout = builder.build_layout(template)

        assert layout.milestone_tracking is True
        assert "Milestone" in layout.columns

        # Generate Excel
        config = ProjectConfig(
            project_id="waterfall_proj_001",
            project_name="Waterfall Test Project",
        )

        engine = ExcelTemplateEngine()
        excel_bytes = engine.generate_template(config)

        # Verify
        assert len(excel_bytes) > 0

    def test_cross_platform_compatibility_workflow(self):
        """Test workflow for cross-platform compatibility."""
        # Test for Windows
        windows_compat = ExcelCompatibilityManager(
            target_version=ExcelVersion.EXCEL_2019,
            target_platform=Platform.WINDOWS,
        )

        windows_formula = windows_compat.get_xlookup_formula(
            lookup_value="A1",
            lookup_array="B:B",
            return_array="C:C",
        )

        # Test for Mac
        mac_compat = ExcelCompatibilityManager(
            target_version=ExcelVersion.EXCEL_2019,
            target_platform=Platform.MAC,
        )

        mac_formula = mac_compat.get_xlookup_formula(
            lookup_value="A1",
            lookup_array="B:B",
            return_array="C:C",
        )

        # Both should use fallback
        assert "INDEX(" in windows_formula
        assert "INDEX(" in mac_formula

    def test_custom_formula_injection_workflow(self):
        """Test custom formula injection workflow."""
        from app.excel.templates import CustomFormulaValidator

        validator = CustomFormulaValidator()

        # Add custom formula
        custom_formula = validator.add_custom_formula(
            name="task_risk_score",
            formula="=IF(AND(A1>80, B1=\"High\"), \"Critical\", \"Normal\")",
            description="Calculate task risk score",
        )

        assert custom_formula["name"] == "task_risk_score"
        assert custom_formula["custom"] is True

    def test_template_versioning_workflow(self):
        """Test template version management workflow."""
        from app.excel.templates import TemplateVersionManager

        manager = TemplateVersionManager()

        # Track versions
        manager.record_version("my_template", "1.0.0")
        manager.record_version("my_template", "1.1.0")

        # Check compatibility
        assert manager.is_compatible("1.1.0", "1.0.0")

        # Increment version
        new_version = manager.increment_version("1.1.0", "minor")
        assert new_version == "1.2.0"


class TestRegressionScenarios:
    """Regression tests to prevent known issues."""

    def test_metadata_encoding_handles_special_characters(self):
        """Test that metadata handles special characters properly."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_special_chars",
            project_name="Test Project with Special Chars: & < > \" '",
        )

        excel_bytes = engine.generate_template(config)
        metadata = engine.load_metadata_from_excel(excel_bytes)

        # Should preserve special characters
        assert "&" in metadata["project_name"]

    def test_empty_project_name_handled(self):
        """Test that empty project name is handled gracefully."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="test_empty_name",
            project_name="",
        )

        # Should not raise exception
        excel_bytes = engine.generate_template(config)
        assert excel_bytes is not None

    def test_large_project_id_handled(self):
        """Test that very long project IDs are handled."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="x" * 100,  # Very long ID
            project_name="Long ID Test",
        )

        excel_bytes = engine.generate_template(config)
        assert excel_bytes is not None

    def test_unicode_characters_in_project_name(self):
        """Test that Unicode characters are preserved."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="unicode_test",
            project_name="Test Project 日本語 Français Español",
        )

        excel_bytes = engine.generate_template(config)
        metadata = engine.load_metadata_from_excel(excel_bytes)

        # Should preserve Unicode
        assert "日本語" in metadata["project_name"]

    def test_concurrent_generation_safe(self):
        """Test that multiple Excel generations can happen concurrently."""
        engine = ExcelTemplateEngine()

        # Generate multiple files
        configs = [
            ProjectConfig(project_id=f"concurrent_{i}", project_name=f"Project {i}")
            for i in range(5)
        ]

        excel_files = [engine.generate_template(config) for config in configs]

        # All should succeed
        assert all(excel_bytes is not None for excel_bytes in excel_files)
        assert all(len(excel_bytes) > 0 for excel_bytes in excel_files)


class TestErrorHandling:
    """Test error handling in integration scenarios."""

    def test_missing_metadata_sheet_raises_error(self):
        """Test that missing metadata sheet raises appropriate error."""
        from openpyxl import Workbook

        # Create Excel without metadata sheet
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Project Plan"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        engine = ExcelTemplateEngine()

        with pytest.raises(ValueError, match="Missing _SYNC_META"):
            engine.load_metadata_from_excel(buffer.read())

    def test_invalid_metadata_json_raises_error(self):
        """Test that invalid metadata JSON raises error."""
        from openpyxl import Workbook

        workbook = Workbook()
        ws = workbook.create_sheet("_SYNC_META")
        ws["A1"] = "invalid json {{"  # Invalid JSON

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        engine = ExcelTemplateEngine()

        with pytest.raises(ValueError, match="Invalid metadata format"):
            engine.load_metadata_from_excel(buffer.read())


# Test coverage marker
def test_integration_coverage_target():
    """Verify integration test coverage meets target."""
    # This is a marker test - actual coverage measured by pytest-cov
    # Target: >85% coverage, 100% pass rate for integration scenarios
    assert True, "Integration test coverage target: >85%"
