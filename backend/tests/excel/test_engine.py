"""Tests for Excel template engine."""

import json
import pytest
from io import BytesIO
from openpyxl import load_workbook

from app.excel.engine import ExcelTemplateEngine, ProjectConfig


class TestProjectConfig:
    """Test ProjectConfig class."""

    def test_create_basic_config(self):
        """Test creating a basic project configuration."""
        config = ProjectConfig(
            project_id="test_123",
            project_name="Test Project",
        )

        assert config.project_id == "test_123"
        assert config.project_name == "Test Project"
        assert config.sprint_pattern == "YY.Q.#"
        assert config.features == {}
        assert config.metadata == {}

    def test_create_config_with_features(self):
        """Test creating configuration with features."""
        config = ProjectConfig(
            project_id="test_123",
            project_name="Test Project",
            sprint_pattern="PI-N.Sprint-M",
            features={"monte_carlo": True, "resource_leveling": False},
            metadata={"created_by": "test_user"},
        )

        assert config.sprint_pattern == "PI-N.Sprint-M"
        assert config.features["monte_carlo"] is True
        assert config.features["resource_leveling"] is False
        assert config.metadata["created_by"] == "test_user"


class TestExcelTemplateEngine:
    """Test ExcelTemplateEngine class."""

    @pytest.fixture
    def engine(self):
        """Create engine instance for testing."""
        return ExcelTemplateEngine()

    @pytest.fixture
    def basic_config(self):
        """Create basic project configuration for testing."""
        return ProjectConfig(
            project_id="test_proj_001",
            project_name="Test Project",
            sprint_pattern="YY.Q.#",
        )

    def test_engine_initialization(self, engine):
        """Test engine initializes correctly."""
        assert engine is not None
        assert engine.formula_templates is not None
        assert isinstance(engine.components, list)

    def test_generate_template_returns_bytes(self, engine, basic_config):
        """Test template generation returns bytes."""
        result = engine.generate_template(basic_config)

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_template_creates_valid_excel(self, engine, basic_config):
        """Test generated template is a valid Excel file."""
        excel_bytes = engine.generate_template(basic_config)

        # Should be able to load as workbook
        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        assert workbook is not None
        assert len(workbook.sheetnames) > 0

    def test_generated_excel_has_main_worksheet(self, engine, basic_config):
        """Test generated Excel has main project plan worksheet."""
        excel_bytes = engine.generate_template(basic_config)

        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        assert "Project Plan" in workbook.sheetnames

        ws = workbook["Project Plan"]
        # Check header row exists
        assert ws.cell(1, 1).value == "Task ID"
        assert ws.cell(1, 2).value == "Task Name"

    def test_generated_excel_has_sync_metadata(self, engine, basic_config):
        """Test generated Excel has hidden sync metadata worksheet."""
        excel_bytes = engine.generate_template(basic_config)

        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        assert "_SYNC_META" in workbook.sheetnames

        ws = workbook["_SYNC_META"]
        # Check it's hidden
        assert ws.sheet_state == "hidden"

        # Check metadata content
        metadata_json = ws["A1"].value
        assert metadata_json is not None

        metadata = json.loads(metadata_json)
        assert metadata["project_id"] == "test_proj_001"
        assert metadata["project_name"] == "Test Project"
        assert "version" in metadata
        assert "generated_at" in metadata
        assert "checksum" in metadata

    def test_load_metadata_from_excel(self, engine, basic_config):
        """Test loading metadata from generated Excel file."""
        excel_bytes = engine.generate_template(basic_config)

        metadata = engine.load_metadata_from_excel(excel_bytes)

        assert metadata["project_id"] == "test_proj_001"
        assert metadata["project_name"] == "Test Project"
        assert metadata["sprint_pattern"] == "YY.Q.#"
        assert "checksum" in metadata

    def test_load_metadata_from_invalid_excel_raises_error(self, engine):
        """Test loading metadata from non-SprintForge Excel raises error."""
        # Create a simple Excel file without metadata
        workbook = load_workbook()
        workbook.create_sheet("Sheet1")

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        excel_bytes = buffer.read()

        with pytest.raises(ValueError, match="Missing _SYNC_META worksheet"):
            engine.load_metadata_from_excel(excel_bytes)

    def test_main_worksheet_has_headers(self, engine, basic_config):
        """Test main worksheet has all expected headers."""
        excel_bytes = engine.generate_template(basic_config)

        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)
        ws = workbook["Project Plan"]

        expected_headers = [
            "Task ID",
            "Task Name",
            "Duration (days)",
            "Start Date",
            "End Date",
            "Dependencies",
            "Sprint",
            "Status",
            "Owner",
        ]

        for col_idx, expected in enumerate(expected_headers, start=1):
            actual = ws.cell(1, col_idx).value
            assert actual == expected, f"Column {col_idx} header mismatch"

    def test_main_worksheet_has_frozen_panes(self, engine, basic_config):
        """Test main worksheet has frozen header row."""
        excel_bytes = engine.generate_template(basic_config)

        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)
        ws = workbook["Project Plan"]

        assert ws.freeze_panes == "A2"

    def test_main_worksheet_has_sample_task(self, engine, basic_config):
        """Test main worksheet includes a sample task."""
        excel_bytes = engine.generate_template(basic_config)

        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)
        ws = workbook["Project Plan"]

        # Check sample task exists
        assert ws.cell(2, 1).value == "T001"
        assert ws.cell(2, 2).value == "Sample Task"
        assert ws.cell(2, 3).value == 5

    def test_checksum_is_consistent(self, engine, basic_config):
        """Test checksum calculation is consistent."""
        excel_bytes_1 = engine.generate_template(basic_config)
        metadata_1 = engine.load_metadata_from_excel(excel_bytes_1)

        excel_bytes_2 = engine.generate_template(basic_config)
        metadata_2 = engine.load_metadata_from_excel(excel_bytes_2)

        # Same config should produce same checksum
        assert metadata_1["checksum"] == metadata_2["checksum"]

    def test_checksum_changes_with_different_config(self, engine):
        """Test checksum changes when configuration changes."""
        config1 = ProjectConfig(
            project_id="proj_1", project_name="Project 1", sprint_pattern="YY.Q.#"
        )
        config2 = ProjectConfig(
            project_id="proj_2", project_name="Project 2", sprint_pattern="PI-N.S-M"
        )

        excel_bytes_1 = engine.generate_template(config1)
        excel_bytes_2 = engine.generate_template(config2)

        metadata_1 = engine.load_metadata_from_excel(excel_bytes_1)
        metadata_2 = engine.load_metadata_from_excel(excel_bytes_2)

        # Different configs should produce different checksums
        assert metadata_1["checksum"] != metadata_2["checksum"]


class TestExcelGeneration:
    """Integration tests for complete Excel generation workflow."""

    def test_end_to_end_generation_and_loading(self):
        """Test complete workflow: generate → save → load → verify."""
        engine = ExcelTemplateEngine()
        config = ProjectConfig(
            project_id="e2e_test",
            project_name="End-to-End Test",
            sprint_pattern="YY.Q.#",
            features={"gantt_chart": True},
        )

        # Generate
        excel_bytes = engine.generate_template(config)
        assert len(excel_bytes) > 0

        # Load
        metadata = engine.load_metadata_from_excel(excel_bytes)
        assert metadata["project_id"] == "e2e_test"

        # Verify workbook structure
        buffer = BytesIO(excel_bytes)
        workbook = load_workbook(buffer)

        assert "Project Plan" in workbook.sheetnames
        assert "_SYNC_META" in workbook.sheetnames
        assert len(workbook.sheetnames) == 2

    def test_multiple_projects_generate_correctly(self):
        """Test engine can generate multiple projects without interference."""
        engine = ExcelTemplateEngine()

        projects = [
            ProjectConfig("p1", "Project Alpha", "YY.Q.#"),
            ProjectConfig("p2", "Project Beta", "PI-N.S-M"),
            ProjectConfig("p3", "Project Gamma", "YYYY-WW"),
        ]

        generated_files = []

        for config in projects:
            excel_bytes = engine.generate_template(config)
            metadata = engine.load_metadata_from_excel(excel_bytes)

            assert metadata["project_id"] == config.project_id
            assert metadata["project_name"] == config.project_name
            assert metadata["sprint_pattern"] == config.sprint_pattern

            generated_files.append(excel_bytes)

        # Verify all files are different
        assert len(generated_files) == 3
        assert generated_files[0] != generated_files[1]
        assert generated_files[1] != generated_files[2]
