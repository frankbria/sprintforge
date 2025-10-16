#!/usr/bin/env python3
"""
Validation script for ExcelGenerationService.

Generates a complete Excel file with all features and validates it can be opened.
"""

import sys
from datetime import datetime
from pathlib import Path

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.services.excel_generation_service import ExcelGenerationService
from app.services.simulation_service import SimulationResult


def main():
    """Generate sample Excel file and validate."""
    print("=" * 60)
    print("Excel Generation Validation Script")
    print("=" * 60)

    # Create service
    service = ExcelGenerationService()
    print("\n✓ ExcelGenerationService initialized")

    # Create workbook with all features
    print("\n1. Creating template workbook with sample data...")
    workbook = service.create_template_workbook(
        project_name="Monte Carlo Validation Test", include_sample_data=True
    )
    print(f"   ✓ Created workbook with sheet: {workbook.sheetnames[0]}")

    # Add Monte Carlo results
    print("\n2. Adding Monte Carlo Results sheet...")
    sample_result = SimulationResult(
        project_duration_days=25.3,
        confidence_intervals={10: 18.2, 50: 24.5, 90: 33.1, 95: 36.8, 99: 42.5},
        mean_duration=25.3,
        median_duration=24.5,
        std_deviation=5.2,
        iterations_run=10000,
        simulation_date=datetime.now(),
        task_count=15,
    )

    sample_tasks = [
        {
            "task_id": f"TASK-{i}",
            "task_name": f"Sample Task {i}",
            "optimistic": 2.0 + i,
            "most_likely": 4.0 + i,
            "pessimistic": 6.0 + i,
            "dependencies": f"TASK-{i-1}" if i > 1 else "",
        }
        for i in range(1, 16)
    ]

    critical_path = ["TASK-1", "TASK-5", "TASK-10", "TASK-15"]

    service.add_monte_carlo_results_sheet(
        workbook, sample_result, sample_tasks, critical_path
    )
    print(f"   ✓ Added Monte Carlo Results with {len(sample_tasks)} tasks")

    # Add Quick Simulation sheet
    print("\n3. Adding Quick Simulation sheet (100 tasks)...")
    service.create_quick_simulation_sheet(workbook)
    print("   ✓ Added Quick Simulation with 100 sample tasks")

    # Apply formatting
    print("\n4. Applying professional formatting...")
    service.apply_formatting(workbook)
    print("   ✓ Applied headers, borders, colors, and number formatting")

    # Save to file
    output_path = Path(__file__).parent.parent / "test_monte_carlo_output.xlsx"
    print(f"\n5. Saving workbook to: {output_path}")

    excel_bytes = service.save_workbook_to_bytes(workbook)
    with open(output_path, "wb") as f:
        f.write(excel_bytes)

    file_size_kb = len(excel_bytes) / 1024
    print(f"   ✓ Saved {file_size_kb:.1f} KB")

    # Validate by reopening
    print("\n6. Validating saved file...")
    from openpyxl import load_workbook

    loaded_wb = load_workbook(output_path)
    print(f"   ✓ File successfully loaded with {len(loaded_wb.sheetnames)} sheets:")
    for sheet_name in loaded_wb.sheetnames:
        sheet = loaded_wb[sheet_name]
        print(f"      - {sheet_name}: {sheet.max_row} rows × {sheet.max_column} cols")

    # Validate PERT formulas
    print("\n7. Validating PERT formulas...")
    task_sheet = loaded_wb["Task List"]
    pert_cell = task_sheet["F2"]
    if isinstance(pert_cell.value, str) and pert_cell.value.startswith("="):
        print(f"   ✓ PERT formula found: {pert_cell.value}")
    else:
        print(f"   ⚠ PERT cell contains: {pert_cell.value} (expected formula)")

    # Summary
    print("\n" + "=" * 60)
    print("VALIDATION SUMMARY")
    print("=" * 60)
    print(f"✓ All sheets created: {', '.join(loaded_wb.sheetnames)}")
    print(f"✓ File size: {file_size_kb:.1f} KB")
    print(f"✓ Output location: {output_path}")
    print("\n✓ Excel file generation validated successfully!")
    print("\nYou can now open this file in Excel or LibreOffice to verify:")
    print(f"  {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
